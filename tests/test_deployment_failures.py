"""TDD tests for §11.11 Deployment Failures — Excel sheet and dashboard route."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch
from datetime import datetime, timezone

import pytest

from cloudopt.models import (
    CollectionMetadata,
    DeploymentFailureEntry,
    VmInventory,
    VmMetrics,
    VmRecommendation,
)

SUB_ID = "a1b2c3d4-ef56-7890-abcd-ef1234567890"


def _metadata() -> CollectionMetadata:
    from cloudopt.models import CollectionThresholds
    return CollectionMetadata(
        run_date="2026-05-12",
        tool_version="0.0.1",
        subscriptions_scanned=[],
        metrics_period_days=30,
        total_vm_count=0,
        thresholds=CollectionThresholds(),
    )


def _failure(
    error_class: str = "allocation",
    resource_name: str = "vm1",
    resource_type: str = "microsoft.compute/virtualmachines",
) -> DeploymentFailureEntry:
    return DeploymentFailureEntry(
        resource_id=f"/subscriptions/{SUB_ID}/resourceGroups/rg/providers/Microsoft.Compute/virtualMachines/{resource_name}",
        resource_name=resource_name,
        resource_type=resource_type,
        subscription_id=SUB_ID,
        resource_group="rg",
        region="eastus",
        error_class=error_class,
        operation_name="Microsoft.Compute/virtualMachines/write",
        status_message="AllocationFailed: No capacity",
        timestamp="2026-03-01T00:00:00Z",
    )


# ---------------------------------------------------------------------------
# Excel sheet tests
# ---------------------------------------------------------------------------

class TestDeploymentFailuresSheet:
    def test_sheet_created_when_data_present(self, tmp_path: Path):
        """write_workbook creates 'Deployment Failures' sheet."""
        import openpyxl
        from cloudopt.export.excel import write_workbook

        path = tmp_path / "out.xlsx"
        write_workbook(
            vms=[],
            metrics=[],
            recommendations=[],
            metadata=_metadata(),
            path=path,
            deployment_failures=[_failure()],
        )
        wb = openpyxl.load_workbook(path)
        assert "Deployment Failures" in wb.sheetnames

    def test_sheet_created_when_no_data(self, tmp_path: Path):
        """'Deployment Failures' sheet is still created when list is empty."""
        import openpyxl
        from cloudopt.export.excel import write_workbook

        path = tmp_path / "out.xlsx"
        write_workbook(
            vms=[],
            metrics=[],
            recommendations=[],
            metadata=_metadata(),
            path=path,
            deployment_failures=[],
        )
        wb = openpyxl.load_workbook(path)
        assert "Deployment Failures" in wb.sheetnames

    def test_sheet_headers_no_cost_columns(self, tmp_path: Path):
        """Headers contain expected columns and no $ / cost fields."""
        import openpyxl
        from cloudopt.export.excel import write_workbook

        path = tmp_path / "out.xlsx"
        write_workbook(
            vms=[],
            metrics=[],
            recommendations=[],
            metadata=_metadata(),
            path=path,
            deployment_failures=[_failure()],
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Deployment Failures"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

        assert "Error Class" in headers
        assert "Resource Name" in headers
        assert "Resource Type" in headers
        assert "Timestamp" in headers
        # No cost columns
        assert not any(h and "$" in h for h in headers)

    def test_sheet_data_row_written(self, tmp_path: Path):
        """A failure entry is written as a data row."""
        import openpyxl
        from cloudopt.export.excel import write_workbook

        path = tmp_path / "out.xlsx"
        write_workbook(
            vms=[],
            metrics=[],
            recommendations=[],
            metadata=_metadata(),
            path=path,
            deployment_failures=[_failure("quota", "vm2")],
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Deployment Failures"]
        # Row 1 = header; row 2 = first data row
        row_values = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        assert any(v == "quota" for v in row_values)
        assert any(v == "vm2" for v in row_values)


# ---------------------------------------------------------------------------
# Dashboard route tests
# ---------------------------------------------------------------------------

class TestDeploymentFailuresDashboardRoute:
    def test_api_returns_list(self):
        """GET /api/deployment-failures returns a list."""
        from fastapi.testclient import TestClient
        from cloudopt.dashboard.app import create_app

        with patch("cloudopt.dashboard.app._load"):
            app = create_app(MagicMock())

        from cloudopt.dashboard.app import _DATA
        _DATA["deployment_failures"] = [
            _failure("allocation"),
            _failure("quota", "vm2"),
        ]

        client = TestClient(app)
        resp = client.get("/api/deployment-failures")
        assert resp.status_code == 200
        data = resp.json()
        assert isinstance(data, list)
        assert len(data) == 2

    def test_api_fields_no_cost(self):
        """Response objects have error_class, resource_name, timestamp and no $ fields."""
        from fastapi.testclient import TestClient
        from cloudopt.dashboard.app import create_app

        with patch("cloudopt.dashboard.app._load"):
            app = create_app(MagicMock())

        from cloudopt.dashboard.app import _DATA
        _DATA["deployment_failures"] = [_failure("image")]

        client = TestClient(app)
        resp = client.get("/api/deployment-failures")
        item = resp.json()[0]
        assert "error_class" in item
        assert "resource_name" in item
        assert "timestamp" in item
        assert not any("$" in k for k in item.keys())

    def test_filter_by_error_class(self):
        """?error_class=allocation filters results."""
        from fastapi.testclient import TestClient
        from cloudopt.dashboard.app import create_app

        with patch("cloudopt.dashboard.app._load"):
            app = create_app(MagicMock())

        from cloudopt.dashboard.app import _DATA
        _DATA["deployment_failures"] = [
            _failure("allocation", "vm1"),
            _failure("quota", "vm2"),
            _failure("image", "vm3"),
        ]

        client = TestClient(app)
        resp = client.get("/api/deployment-failures?error_class=allocation")
        data = resp.json()
        assert all(d["error_class"] == "allocation" for d in data)
        assert len(data) == 1
