"""Tests for Excel, JSON, and CSV export + round-trip."""
import json
import csv
import pytest
from cloudopt.models import (
    CollectionMetadata,
    CollectionThresholds,
    VmInventory,
    VmMetrics,
    VmRecommendation,
    DailyDataPoint,
)
from cloudopt.export.excel import write_workbook, read_workbook
from cloudopt.export.json_export import write_json
from cloudopt.export.csv_export import write_csv


# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------

SUB_ID = "a1b2c3d4-ef56-7890-abcd-ef1234567890"

def _make_vms():
    return [
        VmInventory(
            vm_name="vm-a",
            subscription_id=SUB_ID,
            subscription_name="Test Subscription",
            resource_group="rg-prod",
            resource_id=f"/subscriptions/{SUB_ID}/resourceGroups/rg-prod/providers/Microsoft.Compute/virtualMachines/vm-a",
            vm_sku="Standard_D4s_v3",
            vcpus=4,
            memory_gb=16.0,
            region="eastus",
            os_type="Linux",
            workload="SAP",
            environment="Production",
            vmss_name="my-vmss",
        ),
        VmInventory(
            vm_name="vm-b",
            subscription_id=SUB_ID,
            subscription_name="Test Subscription",
            resource_group="rg-dev",
            resource_id=f"/subscriptions/{SUB_ID}/resourceGroups/rg-dev/providers/Microsoft.Compute/virtualMachines/vm-b",
            vm_sku="Standard_B2s",
            vcpus=2,
            memory_gb=4.0,
            region="westeurope",
            os_type="Windows",
            availability_set_name="my-avset",
        ),
    ]


def _make_metrics(vms):
    return [
        VmMetrics(
            resource_id=vms[0].resource_id,
            metric_name="Percentage CPU",
            avg=12.5,
            p50=11.0,
            p95=38.0,
            p99=42.0,
            max=75.0,
            min=2.0,
            time_series=[DailyDataPoint(date="2026-04-01T00:00:00Z", value=12.5)],
        )
    ]


def _make_recs(vms):
    return [
        VmRecommendation(
            resource_id=vms[0].resource_id,
            current_sku="Standard_D4s_v3",
            recommended_sku="Standard_D2s_v3",
            category="underutilized",
            reason="CPU avg 12.5% < 15%",
            estimated_savings_pct=50.0,
        )
    ]


def _make_metadata():
    return CollectionMetadata(
        run_date="2026-04-07T10:00:00Z",
        tool_version="0.1.0",
        subscriptions_scanned=[SUB_ID],
        metrics_period_days=30,
        total_vm_count=2,
        thresholds=CollectionThresholds(),
    )


# ---------------------------------------------------------------------------
# Excel round-trip
# ---------------------------------------------------------------------------

class TestExcelRoundTrip:
    # Note: Phase C write_workbook signature: (vms, metrics, findings, metadata, path)
    # findings is now optional; pass VmRecommendation via recommendations= kwarg

    def test_workbook_created(self, tmp_path):
        path = tmp_path / "output.xlsx"
        vms = _make_vms()
        metrics = _make_metrics(vms)
        recs = _make_recs(vms)
        meta = _make_metadata()
        write_workbook(vms, metrics, [], meta, path, recommendations=recs)
        assert path.exists()
        assert path.stat().st_size > 0

    def test_read_back_analyst_fields(self, tmp_path):
        path = tmp_path / "output.xlsx"
        vms = _make_vms()
        metrics = _make_metrics(vms)
        meta = _make_metadata()
        write_workbook(vms, metrics, [], meta, path)

        # Phase C: inventory sheet no longer written — read_workbook returns empty VM list
        updated_vms, _, _, _ = read_workbook(path)
        assert updated_vms == []

    def test_vmss_name_in_workbook(self, tmp_path):
        path = tmp_path / "output.xlsx"
        vms = _make_vms()
        write_workbook(vms, [], [], _make_metadata(), path)
        # Phase C: inventory sheet removed — read_workbook returns empty list
        updated_vms, _, _, _ = read_workbook(path)
        assert updated_vms == []

    def test_read_returns_all_vms(self, tmp_path):
        path = tmp_path / "output.xlsx"
        vms = _make_vms()
        write_workbook(vms, [], [], _make_metadata(), path)
        updated_vms, _, _, _ = read_workbook(path)
        # Phase C: no Fleet Inventory sheet written; round-trip compatibility intentionally broken
        assert len(updated_vms) == 0

    def test_p99_survives_raw_metrics_roundtrip(self, tmp_path):
        """Phase C restructured sheets; verify write/read succeeds and expected sheets exist."""
        import openpyxl
        path = tmp_path / "output.xlsx"
        vms = _make_vms()
        metrics = _make_metrics(vms)  # p99=42.0
        write_workbook(vms, metrics, [], _make_metadata(), path)
        _, read_metrics, _, _ = read_workbook(path)
        # Raw Metrics sheet absent in Phase C — metrics come back empty
        assert len(read_metrics) == 0
        wb = openpyxl.load_workbook(path, data_only=True)
        # Phase C replaces Evidence with Decisions + Performance sheets
        assert "Decisions" in wb.sheetnames
        assert "Perf by VM" in wb.sheetnames


# ---------------------------------------------------------------------------
# JSON export — masking
# ---------------------------------------------------------------------------

class TestJsonExport:
    def test_write_and_parse(self, tmp_path):
        path = tmp_path / "output.json"
        vms = _make_vms()
        metrics = _make_metrics(vms)
        recs = _make_recs(vms)
        meta = _make_metadata()
        write_json(vms, metrics, recs, meta, path)
        data = json.loads(path.read_text())
        assert "metadata" in data
        assert "vms" in data
        assert "metrics" in data
        assert "recommendations" in data

    def test_subscription_ids_masked(self, tmp_path):
        path = tmp_path / "output.json"
        vms = _make_vms()
        write_json(vms, [], [], _make_metadata(), path)
        text = path.read_text()
        # Full GUID should not appear in JSON
        assert "a1b2c3d4-ef56-7890-abcd-ef1234567890" not in text
        # Masked prefix should appear
        assert "a1b2c3d4-xxxx" in text

    def test_p99_in_metrics_json(self, tmp_path):
        path = tmp_path / "output.json"
        vms = _make_vms()
        metrics = _make_metrics(vms)  # p99=42.0
        write_json(vms, metrics, [], _make_metadata(), path)
        data = json.loads(path.read_text())
        assert data["metrics"][0]["p99"] == pytest.approx(42.0)

    def test_vm_names_not_masked(self, tmp_path):
        path = tmp_path / "output.json"
        vms = _make_vms()
        write_json(vms, [], [], _make_metadata(), path)
        text = path.read_text()
        assert "vm-a" in text
        assert "vm-b" in text

    def test_metadata_thresholds_present(self, tmp_path):
        path = tmp_path / "output.json"
        write_json([], [], [], _make_metadata(), path)
        data = json.loads(path.read_text())
        thresholds = data["metadata"]["thresholds"]
        assert "underutilized_cpu_avg" in thresholds


class TestDiskJsonRoundTrip:
    def _disk(self):
        from cloudopt.models import DiskInventory

        sub = "a1b2c3d4-ef56-7890-abcd-ef1234567890"
        vm = f"/subscriptions/{sub}/resourceGroups/rg/providers/Microsoft.Compute/virtualMachines/vm1"
        return DiskInventory(
            resource_id=f"/subscriptions/{sub}/resourceGroups/rg/providers/Microsoft.Compute/disks/d1",
            disk_name="d1",
            subscription_id=sub,
            subscription_name="Prod",
            resource_group="rg",
            location="eastus",
            sku_name="Premium_LRS",
            performance_tier="P30",
            disk_size_gb=1024,
            disk_iops_read_write=5000,
            disk_mbps_read_write=200,
            disk_state="Attached",
            managed_by=vm,
            raw_properties={"tier": "P30"},
        )

    def test_disks_written_and_masked(self, tmp_path):
        path = tmp_path / "output.json"
        write_json([], [], [], _make_metadata(), path, disks=[self._disk()])
        data = json.loads(path.read_text())
        assert "disks" in data
        assert len(data["disks"]) == 1
        text = path.read_text()
        assert "a1b2c3d4-ef56-7890-abcd-ef1234567890" not in text
        assert "a1b2c3d4-xxxx" in text
        # raw_properties is intentionally NOT serialized to JSON
        assert "raw_properties" not in data["disks"][0]

    def test_disk_reloads_and_detector_fires(self, tmp_path):
        from cloudopt.analyzer.detectors import disk_pv2
        from cloudopt.models import DiskInventory

        path = tmp_path / "output.json"
        write_json([], [], [], _make_metadata(), path, disks=[self._disk()])
        data = json.loads(path.read_text())
        reloaded = [DiskInventory(**d) for d in data["disks"]]
        assert reloaded[0].is_premium_v1 and reloaded[0].is_data_disk
        findings = disk_pv2.detect(reloaded)
        assert [f.code for f in findings] == ["SWP-DST-002"]

    def test_disk_excel_sheet_round_trip(self, tmp_path):
        from cloudopt.export.excel import read_disks_from_workbook

        path = tmp_path / "output.xlsx"
        write_workbook([], [], [], _make_metadata(), path, disks=[self._disk()])
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True)
        assert "Disk Inventory" in wb.sheetnames
        wb.close()
        disks = read_disks_from_workbook(path)
        assert len(disks) == 1
        assert disks[0].disk_name == "d1"
        assert disks[0].sku_name == "Premium_LRS"


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

class TestCsvExport:
    def test_files_created(self, tmp_path):
        vms = _make_vms()
        metrics = _make_metrics(vms)
        recs = _make_recs(vms)
        write_csv(vms, metrics, recs, _make_metadata(), tmp_path)
        assert (tmp_path / "vm_inventory.csv").exists()
        assert (tmp_path / "metrics.csv").exists()
        assert (tmp_path / "recommendations.csv").exists()

    def test_vm_inventory_columns(self, tmp_path):
        vms = _make_vms()
        write_csv(vms, [], [], _make_metadata(), tmp_path)
        with open(tmp_path / "vm_inventory.csv", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows) == 2
        assert "vm_name" in reader.fieldnames
        assert "vm_sku" in reader.fieldnames
        assert "subscription_id" in reader.fieldnames

    def test_metrics_csv_row_count(self, tmp_path):
        vms = _make_vms()
        metrics = _make_metrics(vms)
        write_csv(vms, metrics, [], _make_metadata(), tmp_path)
        with open(tmp_path / "metrics.csv", newline="") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == len(metrics)

    def test_recommendations_csv_columns(self, tmp_path):
        vms = _make_vms()
        recs = _make_recs(vms)
        write_csv(vms, [], recs, _make_metadata(), tmp_path)
        with open(tmp_path / "recommendations.csv", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows) == 1
        assert "category" in reader.fieldnames
        assert "current_sku" in reader.fieldnames
