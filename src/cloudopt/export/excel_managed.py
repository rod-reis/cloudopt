"""Managed-service sheet writers for the cloudopt Excel workbook (SPEC §7.3).

Sheets produced (one per service type):
  10. AKS
  11. AVD
  12. Databricks
  13. Azure Batch
  14. AML
  15. ARO
  16. HDInsight

All sheets share the same column schema.  The sheet name is the service type's
human-readable value (e.g. "AKS", "AVD", etc.).
"""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from cloudopt.models import ManagedComputeGroupRow, ParentServiceType

from cloudopt.export.excel import (
    _ALT_FILL,
    _THIN_BORDER,
    _write_header,
    _add_table,
    _colour_util,
)

# Re-use the 39-column guest metric definition from excel_perf
from cloudopt.export.excel_perf import _GUEST_COLS

# ---------------------------------------------------------------------------
# SPEC §7.3 — static columns for managed-service sheets
# ---------------------------------------------------------------------------
_MANAGED_STATIC: list[tuple[str, str, int]] = [
    ("Parent Service Type",  "parent_service_type",  18),
    ("Parent Service Name",  "parent_service_name",  28),
    ("Parent Service ID",    "parent_service_id",    50),
    ("Pool/NodePool Name",   "parent_pool_name",     22),
    ("VMSS Name",            "vmss_name",            22),
    ("VMSS ID",              "vmss_id",              50),
    ("VM SKU",               "vm_sku",               22),
    ("Instance Count",       "instance_count",       14),
    ("Subscription",         "subscription_name",    26),
    ("Resource Group",       "resource_group",       24),
    ("Region",               "region",               16),
    ("OS Type",              "os_type",              12),
    ("OS Image",             "os_image",             24),
    ("Zones",                "zones",                14),
    ("vCPUs",                "vcpus",                8),
    ("Mem (GB)",             "memory_gb",            10),
    ("Tags",                 "tags",              28),
]

_MANAGED_PLATFORM: list[tuple[str, str]] = [
    ("Avg CPU %",    "avg_cpu_pct"),
    ("P95 CPU %",    "p95_cpu_pct"),
    ("P99 CPU %",    "p99_cpu_pct"),
    ("Max CPU %",    "max_cpu_pct"),
    ("Min CPU %",    "min_cpu_pct"),
    ("Avg Mem %",    "avg_mem_pct"),
]

_MANAGED_TRAILING: list[tuple[str, str, int]] = [
    ("Has OS Data",    "has_os_data",    12),
    ("Sources Used",   "sources_used",   20),
    ("Days Observed",  "days_observed",  14),
    ("Coverage %",     "coverage_pct",   12),
]


def sheet_managed_service(
    wb: Workbook,
    svc_type: ParentServiceType,
    groups: list[ManagedComputeGroupRow],
) -> None:
    """Write one managed-service sheet for *svc_type* with rows from *groups*."""
    sheet_name = svc_type.value  # "AKS", "AVD", "Databricks", etc.
    ws = wb.create_sheet(sheet_name)

    static_hdrs = [c[0] for c in _MANAGED_STATIC]
    platform_hdrs = [c[0] for c in _MANAGED_PLATFORM]
    guest_hdrs = [label for label, _ in _GUEST_COLS]
    trailing_hdrs = [c[0] for c in _MANAGED_TRAILING]
    headers = static_hdrs + platform_hdrs + guest_hdrs + trailing_hdrs
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_idx, grp in enumerate(groups, start=2):
        alt = row_idx % 2 == 0
        row_vals: list[Any] = []

        # Static columns
        for _, field, _ in _MANAGED_STATIC:
            val = getattr(grp, field, None)
            if hasattr(val, "value"):  # Enum → string
                val = val.value
            row_vals.append(val if val is not None else "")

        # Platform metric columns
        for _, field in _MANAGED_PLATFORM:
            row_vals.append(getattr(grp, field, None))

        # Guest metric columns (stored as dict on ManagedComputeGroupRow)
        guest_data = grp.guest_metrics or {}
        for _, field_name in _GUEST_COLS:
            row_vals.append(guest_data.get(field_name))

        # Trailing columns
        for _, field, _ in _MANAGED_TRAILING:
            row_vals.append(getattr(grp, field, None))

        # Write cells
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            if alt:
                cell.fill = _ALT_FILL

        # Colour-code platform CPU columns
        cpu_start = len(_MANAGED_STATIC) + 1
        for cpu_offset in range(5):  # Avg/P95/P99/Max/Min CPU
            col_idx = cpu_start + cpu_offset
            _colour_util(ws.cell(row=row_idx, column=col_idx), row_vals[col_idx - 1])

    # Column widths
    all_widths = (
        [c[2] for c in _MANAGED_STATIC]
        + [13] * len(_MANAGED_PLATFORM)
        + [16] * len(_GUEST_COLS)
        + [c[2] for c in _MANAGED_TRAILING]
    )
    for col_idx, w in enumerate(all_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    safe_name = sheet_name.replace(" ", "").replace("-", "")
    if groups:
        _add_table(ws, f"Tbl{safe_name}", len(groups))
