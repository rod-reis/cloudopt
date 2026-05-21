"""Multi-sheet Excel workbook generation using openpyxl.

22-sheet model (Phase C — SPEC §7):
  1.  Optimizations            ← RECOMMENDATION findings + Advisor SKU changes
  2.  Optimization Candidates  ← CANDIDATE findings requiring customer input
  3.  Summary                  ← KPI block, charts, top-10 READY decisions
  4.  Anomalies                ← platform vs external metric disagreements
  5.  Quota Posture            ← per-subscription quota utilization
  6.  Perf by VM - Standalone  ← per-VM platform + guest metrics (standalone only)
  7.  Perf by VM Group per SKU ← aggregated managed-service group rows
  8.  Perf by VM SKU per Subscription
  9.  Perf by VM SKU per Resource Group
  10. AKS
  11. AVD
  12. Databricks
  13. Azure Batch
  14. AML
  15. ARO
  16. HDInsight
  17. Resource Inventory        ← ARG resource list
  18. Workload Information
  19. Capacity Reservations
  20. Deployment Failures
  21. Run Metadata
  22. App Insights
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.chart import BarChart, PieChart, Reference

from cloudopt.analyzer.taxonomy import Category, Confidence, FindingType, Readiness
from cloudopt.enrichment.schema import EnrichedVmMetrics
from cloudopt.models import (
    AdvisorRecommendation,
    AppInsightsInventory,
    AppInsightsMetrics,
    AzureResource,
    CapacityReservationGroup,
    CollectionMetadata,
    CollectionThresholds,
    DeploymentFailureEntry,
    Finding,
    ManagedComputeGroupRow,
    ParentServiceType,
    QuotaItem,
    SubscriptionZoneMapping,
    VmInventory,
    VmMetrics,
    VmRecommendation,
    WorkloadInfo,
    mask_subscription_id,
    mask_subscription_ids_in_string,
)

# ---------------------------------------------------------------------------
# Colour constants
# ---------------------------------------------------------------------------
_HDR_FILL = PatternFill("solid", fgColor="1F4E79")      # dark blue header
_HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
_ALT_FILL = PatternFill("solid", fgColor="D6E4F0")      # alternating row
_CSA_FILL = PatternFill("solid", fgColor="BDD7EE")      # Analyst-editable column
_GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
_YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
_RED_FILL = PatternFill("solid", fgColor="FFC7CE")

# Readiness tier fills
_READY_FILL = PatternFill("solid", fgColor="C6EFCE")    # green — READY
_LIKELY_FILL = PatternFill("solid", fgColor="FFEB9C")   # yellow — LIKELY
_DISCOVERY_FILL = PatternFill("solid", fgColor="D9D9D9") # grey — DISCOVERY
_INSUFF_FILL = PatternFill("solid", fgColor="FFC7CE")   # red — INSUFFICIENT

# Confidence tier fills
_HIGH_FILL = PatternFill("solid", fgColor="C6EFCE")     # green
_MED_FILL = PatternFill("solid", fgColor="FFEB9C")      # yellow
_LOW_FILL = PatternFill("solid", fgColor="FFC7CE")      # red

# Technical Summary beautification fills
_SECTION_FILL  = PatternFill("solid", fgColor="2E75B6")  # azure blue — section sub-headers
_BAND_FILL     = PatternFill("solid", fgColor="F2F7FC")  # very light blue — KPI row bands
_COL_HDR_FILL  = PatternFill("solid", fgColor="4472C4")  # periwinkle — column sub-headers

_THIN_BORDER = Border(
    left=Side(style="thin", color="B8CCE4"),
    right=Side(style="thin", color="B8CCE4"),
    bottom=Side(style="thin", color="B8CCE4"),
)


# ---------------------------------------------------------------------------
# Public interface
# ---------------------------------------------------------------------------

def write_workbook(
    vms: list[VmInventory],
    metrics: list[VmMetrics],
    findings: list[Finding] | None = None,
    metadata: CollectionMetadata | None = None,
    path: Path | None = None,
    *,
    advisor: list[AdvisorRecommendation] | None = None,
    enriched_metrics: list[EnrichedVmMetrics] | None = None,
    quota_items: list[QuotaItem] | None = None,
    capacity_reservations: list[CapacityReservationGroup] | None = None,
    deployment_failures: list[DeploymentFailureEntry] | None = None,
    resources: list[AzureResource] | None = None,
    workload_info: WorkloadInfo | None = None,
    managed_groups: list[ManagedComputeGroupRow] | None = None,
    app_insights: list[Any] | None = None,
    # Legacy compat params kept as kwargs to avoid breaking callers
    recommendations: list[Any] | None = None,
    quota: list[QuotaItem] | None = None,
    appinsights: list[Any] | None = None,
    appinsights_metrics: list[Any] | None = None,
    zone_mappings: list[Any] | None = None,
) -> None:
    """Write the full Excel workbook to *path* (22-sheet Phase C model)."""
    if path is None:
        raise TypeError("write_workbook() requires a 'path' argument")
    if metadata is None:
        metadata = CollectionMetadata(
            run_date="",
            tool_version="",
            subscriptions_scanned=[],
            metrics_period_days=0,
            total_vm_count=len(vms),
            thresholds=CollectionThresholds(),
        )
    from cloudopt.export.excel_perf import (
        sheet_perf_standalone,
        sheet_perf_group_by_sku,
        sheet_sku_by_subscription,
        sheet_sku_by_resource_group,
    )
    from cloudopt.export.excel_managed import sheet_managed_service

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    metrics_by_vm = _group_metrics(metrics)
    findings_list: list[Finding] = findings or []
    enriched_list: list[EnrichedVmMetrics] = enriched_metrics or []

    # Resolve legacy param aliases
    _quota_items = quota_items or quota or []
    _appinsights = app_insights or appinsights or []
    _appinsights_metrics: list[Any] = appinsights_metrics or []

    # Build VmMetrics dict for new-style callers (keyed by resource_id)
    vm_metrics_dict: dict[str, VmMetrics] = {}
    for m_list in metrics_by_vm.values():
        for m in m_list.values():
            vm_metrics_dict[m.resource_id] = m

    # Sheet 1 — Technical Summary (SPEC: first sheet for immediate context)
    _sheet_technical_summary(wb, findings_list, vms, enriched_list)
    # Sheet 2 — Workload Information
    _wi = workload_info if workload_info is not None else WorkloadInfo()
    _sheet_workload_info(wb, _wi)
    # Sheet 3 — Optimizations (was Decisions)
    _sheet_decisions(wb, findings_list, advisor or [])
    # Sheet 4 — Optimization Candidates (was Discovery Candidates)
    _sheet_discovery_candidates(wb, findings_list)
    # Sheet 5 — Anomalies
    _sheet_anomalies(wb, vms, metrics_by_vm, enriched_list)
    # Sheet 6 — Quota Posture
    _sheet_quota(wb, _quota_items)
    # Sheet 7 — Perf by VM - Standalone
    standalone_vms = [
        v for v in vms
        if v.parent_service_type in (
            ParentServiceType.STANDALONE,
            ParentServiceType.STANDALONE_VMSS,
        )
    ]
    sheet_perf_standalone(wb, standalone_vms, metrics_by_vm, enriched_list)
    # Sheet 8 — Perf by VM Group per SKU
    _groups = managed_groups or []
    sheet_perf_group_by_sku(wb, _groups)
    # Sheet 9 — Perf by VM SKU per Subscription
    sheet_sku_by_subscription(wb, vms, metrics_by_vm)
    # Sheet 10 — Perf by VM SKU per Resource Group
    sheet_sku_by_resource_group(wb, vms, metrics_by_vm)
    # Sheet 11 — Fleet Inventory (VM-only; required by read_workbook → dashboard)
    _sheet_inventory(wb, vms)
    # Sheet 12 — Raw Metrics (required by read_workbook → dashboard)
    _sheet_raw_metrics(wb, metrics)
    # Sheets 13–19 — Managed service sheets
    for svc_type in (
        ParentServiceType.AKS,
        ParentServiceType.AVD,
        ParentServiceType.DATABRICKS,
        ParentServiceType.AZURE_BATCH,
        ParentServiceType.AML,
        ParentServiceType.ARO,
        ParentServiceType.HDINSIGHT,
    ):
        svc_groups = [g for g in _groups if g.parent_service_type == svc_type]
        sheet_managed_service(wb, svc_type, svc_groups)
    # Sheet 20 — Resource Inventory
    _sheet_resources(wb, resources or [])
    # Sheet 21 — Capacity Reservations
    _sheet_capacity_reservations(wb, capacity_reservations or [])
    # Sheet 22 — Deployment Failures
    _sheet_deployment_failures(wb, deployment_failures or [])
    # Sheet 23 — Run Metadata
    _sheet_metadata(wb, metadata)
    # Sheet 24 — App Insights (last per user decision)
    _sheet_appinsights(wb, _appinsights, _appinsights_metrics)

    wb.save(path)


def read_workbook(path: Path) -> tuple[list[VmInventory], list[VmMetrics], list[VmRecommendation], CollectionMetadata]:
    """Read an existing workbook back into model objects.

    Analyst-edited fields (workload, application, environment, criticality, owner,
    custom, manual_override, notes) are preserved from the file.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    vms = _read_inventory_sheet(wb)
    metrics = _read_raw_metrics_sheet(wb)
    recommendations = _read_recommendations_sheet(wb)
    metadata = _read_metadata_sheet(wb)
    return vms, metrics, recommendations, metadata


def read_quota_from_workbook(path: Path) -> list[QuotaItem]:
    """Read the Quota Utilization sheet from an existing workbook."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return _read_quota_sheet(wb)


_RESOURCE_TYPE_TO_SERVICE: dict[str, ParentServiceType] = {
    "microsoft.containerservice/managedclusters": ParentServiceType.AKS,
    "microsoft.desktopvirtualization/hostpools": ParentServiceType.AVD,
    "microsoft.databricks/workspaces": ParentServiceType.DATABRICKS,
    "microsoft.batch/batchaccounts": ParentServiceType.AZURE_BATCH,
    "microsoft.machinelearningservices/workspaces": ParentServiceType.AML,
    "microsoft.redhatopenshift/openshiftclusters": ParentServiceType.ARO,
    "microsoft.hdinsight/clusters": ParentServiceType.HDINSIGHT,
}


def read_vmss_groups_from_workbook(path: Path) -> list[ManagedComputeGroupRow]:
    """Read all VMSS Uniform groups from the 'Perf by VM Group' sheet.

    Includes both standalone VMSS Uniform rows (empty 'Parent ResourceType')
    and managed-service-associated rows (AKS, AVD, Databricks, etc.).
    The parent_service_type is inferred from the 'Parent ResourceType' column.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Perf by VM Group"] if "Perf by VM Group" in wb.sheetnames else None
    if ws is None:
        return []

    # Build header → column-index map from row 1
    col_map: dict[str, int] = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value:
            col_map[str(cell.value)] = cell.column - 1  # zero-based offset

    def _get(row_vals: list, header: str):
        idx = col_map.get(header)
        return row_vals[idx] if idx is not None and idx < len(row_vals) else None

    def _float(v) -> float | None:
        try:
            return float(v) if v is not None else None
        except (TypeError, ValueError):
            return None

    def _int(v) -> int:
        try:
            return int(v) if v is not None else 0
        except (TypeError, ValueError):
            return 0

    results: list[ManagedComputeGroupRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_vals = list(row)
        vmss_name = _get(row_vals, "VMSS Name")
        if not vmss_name:
            continue
        parent_resource_type = _get(row_vals, "Parent ResourceType")
        prt_str = str(parent_resource_type or "").strip().lower()
        svc_type = _RESOURCE_TYPE_TO_SERVICE.get(prt_str, ParentServiceType.STANDALONE_VMSS)
        try:
            results.append(ManagedComputeGroupRow(
                parent_service_type=svc_type,
                parent_resource_type=str(parent_resource_type) if parent_resource_type else None,
                parent_service_name=str(_get(row_vals, "Parent ResourceName") or vmss_name),
                parent_pool_name=str(_get(row_vals, "Parent Pool/Node Group Name") or "") or None,
                vmss_name=str(vmss_name),
                vm_sku=str(_get(row_vals, "VM SKU") or ""),
                instance_count=_int(_get(row_vals, "Instance Count")),
                subscription_name=str(_get(row_vals, "Subscription") or ""),
                subscription_id="",
                resource_group=str(_get(row_vals, "Resource Group") or ""),
                region=str(_get(row_vals, "Region") or ""),
                os_type=str(_get(row_vals, "OS Type") or "") or None,
                avg_cpu_pct=_float(_get(row_vals, "Avg CPU %")),
                p95_cpu_pct=_float(_get(row_vals, "P95 CPU %")),
                p99_cpu_pct=_float(_get(row_vals, "P99 CPU %")),
                max_cpu_pct=_float(_get(row_vals, "Max CPU %")),
                min_cpu_pct=_float(_get(row_vals, "Min CPU %")),
                avg_mem_pct=_float(_get(row_vals, "Avg Mem %")),
            ))
        except Exception:
            pass
    return results


# ---------------------------------------------------------------------------
# Sheet 1: Technical Summary  (Phase B)
# ---------------------------------------------------------------------------

_READINESS_FILL_MAP: dict[str, Any] = {
    "READY": _READY_FILL,
    "LIKELY": _LIKELY_FILL,
    "DISCOVERY": _DISCOVERY_FILL,
    "INSUFFICIENT": _INSUFF_FILL,
}

_CONFIDENCE_FILL_MAP: dict[str, Any] = {
    "HIGH": _HIGH_FILL,
    "MEDIUM": _MED_FILL,
    "LOW": _LOW_FILL,
}


def _sheet_technical_summary(
    wb: Workbook,
    findings: list[Finding],
    vms: list[VmInventory],
    enriched_metrics: list[EnrichedVmMetrics],
) -> None:
    ws = wb.create_sheet("Technical Summary")
    ws.sheet_view.showGridLines = False

    # ── Stats ────────────────────────────────────────────────────────────
    recs       = [f for f in findings if f.finding_type == FindingType.RECOMMENDATION]
    cands      = [f for f in findings if f.finding_type == FindingType.CANDIDATE]
    ready_recs = [f for f in recs if f.readiness == Readiness.READY]
    likely_recs = [f for f in recs if f.readiness == Readiness.LIKELY]
    high_conf  = [f for f in recs if f.confidence == Confidence.HIGH]
    high_pct   = (len(high_conf) / max(len(recs), 1)) * 100
    categories = [c.value for c in Category]

    _thin = Border(
        left=Side(style="thin", color="BDD7EE"),
        right=Side(style="thin", color="BDD7EE"),
        bottom=Side(style="thin", color="BDD7EE"),
    )
    _f_navy  = Font(bold=True, size=12, color="1F4E79")
    _f_green = Font(bold=True, size=12, color="375623")
    _f_amber = Font(bold=True, size=12, color="833C00")
    _f_grey  = Font(bold=True, size=12, color="595959")

    # ── Local helpers ─────────────────────────────────────────────────────
    def _section_header(row: int, text: str, n_cols: int = 5) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = _SECTION_FILL
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.row_dimensions[row].height = 20

    def _col_header(row: int, labels: list) -> None:
        for i, label in enumerate(labels, start=1):
            c = ws.cell(row=row, column=i, value=label)
            c.fill = _COL_HDR_FILL
            c.font = Font(color="FFFFFF", bold=True, size=9)
            c.alignment = Alignment(
                horizontal="center" if i > 1 else "left",
                indent=1 if i == 1 else 0,
                vertical="center",
            )
        ws.row_dimensions[row].height = 18

    def _kpi_row(row: int, label: str, value: Any, desc: str, vfont: Font) -> None:
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font = Font(size=9, color="595959")
        lbl.fill = _BAND_FILL
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lbl.border = _thin

        val = ws.cell(row=row, column=2, value=value)
        val.font = vfont
        val.fill = _BAND_FILL
        val.alignment = Alignment(horizontal="right", vertical="center")
        val.border = _thin

        d = ws.cell(row=row, column=3, value=desc)
        d.font = Font(size=9, color="7F7F7F", italic=True)
        d.fill = _BAND_FILL
        d.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        d.border = _thin
        ws.row_dimensions[row].height = 20

    # ── Row 1: Title banner ───────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    title = ws.cell(row=1, column=1, value="CLOUDOPT  ·  Technical Summary")
    title.font = Font(bold=True, size=16, color="FFFFFF")
    title.fill = _HDR_FILL
    title.alignment = Alignment(vertical="center", horizontal="left", indent=2)
    ws.row_dimensions[1].height = 30

    # ── Row 2: Stats subtitle bar ─────────────────────────────────────────
    ws.merge_cells("A2:J2")
    subtitle = ws.cell(
        row=2, column=1,
        value=(
            f"Fleet: {len(vms)} VMs   ·   {len(recs)} recommendations   ·   "
            f"{len(ready_recs)} READY   ·   {len(likely_recs)} LIKELY   ·   "
            f"{high_pct:.0f}% HIGH confidence"
        ),
    )
    subtitle.font = Font(size=9, color="FFFFFF")
    subtitle.fill = _SECTION_FILL
    subtitle.alignment = Alignment(vertical="center", horizontal="left", indent=2)
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8  # spacer

    # ── Rows 4-11: Fleet Overview KPIs ───────────────────────────────────
    _section_header(4, "  FLEET OVERVIEW")
    _kpi_row(5,  "Fleet Size (VMs)",          len(vms),
             "Total VMs detected across scoped subscriptions",                  _f_navy)
    _kpi_row(6,  "Total Recommendations",     len(recs),
             "Optimization findings generated by the detector pipeline",        _f_navy)
    _kpi_row(7,  "Optimization Candidates",   len(cands),
             "Potential savings requiring additional customer input",            _f_grey)
    _kpi_row(8,  "READY Decisions",           len(ready_recs),
             "Immediately actionable — sufficient evidence collected",          _f_green)
    _kpi_row(9,  "LIKELY Decisions",          len(likely_recs),
             "Likely actionable — validate with customer before committing",    _f_amber)
    _kpi_row(10, "HIGH Confidence %",         f"{high_pct:.1f}%",
             "Fraction of recommendations with strong supporting evidence",
             _f_green if high_pct >= 80 else _f_amber)
    _kpi_row(11, "Monitoring-Enriched VMs",   len(enriched_metrics),
             "VMs with guest OS / application-level metrics available",         _f_navy)

    ws.row_dimensions[12].height = 10  # spacer

    # ── Rows 13-20+: Findings by Category ────────────────────────────────
    _section_header(13, "  FINDINGS BY CATEGORY", n_cols=5)
    _col_header(14, ["Category", "READY", "LIKELY", "DISCOVERY", "Total"])
    cat_data_start = 15
    for i, cat in enumerate(categories):
        row = cat_data_start + i
        n_ready  = sum(1 for f in recs if f.category.value == cat and f.readiness == Readiness.READY)
        n_likely = sum(1 for f in recs if f.category.value == cat and f.readiness == Readiness.LIKELY)
        n_disc   = sum(1 for f in recs if f.category.value == cat and f.readiness == Readiness.DISCOVERY)
        n_total  = n_ready + n_likely + n_disc
        bg = _ALT_FILL if i % 2 == 0 else PatternFill()

        name_c = ws.cell(row=row, column=1, value=cat.capitalize())
        name_c.fill = bg
        name_c.font = Font(size=9)
        name_c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        name_c.border = _thin

        for col_idx, count, active_fill in (
            (2, n_ready,  _READY_FILL),
            (3, n_likely, _LIKELY_FILL),
            (4, n_disc,   _DISCOVERY_FILL),
        ):
            c = ws.cell(row=row, column=col_idx, value=count)
            c.fill = active_fill if count > 0 else bg
            c.font = Font(size=9, bold=count > 0)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin

        tot_c = ws.cell(row=row, column=5, value=n_total)
        tot_c.fill = bg
        tot_c.font = Font(size=9, bold=True)
        tot_c.alignment = Alignment(horizontal="center", vertical="center")
        tot_c.border = _thin
        ws.row_dimensions[row].height = 16

    cat_data_end = cat_data_start + len(categories) - 1
    ws.row_dimensions[cat_data_end + 1].height = 10  # spacer

    # ── Confidence distribution ───────────────────────────────────────────
    conf_hdr_row = cat_data_end + 2
    _section_header(conf_hdr_row, "  CONFIDENCE DISTRIBUTION", n_cols=4)
    _col_header(conf_hdr_row + 1, ["Confidence", "Count", "% of Total"])
    conf_data_start = conf_hdr_row + 2
    conf_items = [
        ("HIGH",   Confidence.HIGH,   _HIGH_FILL, Font(size=9, bold=True, color="375623")),
        ("MEDIUM", Confidence.MEDIUM, _MED_FILL,  Font(size=9, bold=True, color="833C00")),
        ("LOW",    Confidence.LOW,    _LOW_FILL,  Font(size=9, bold=True, color="9C0006")),
    ]
    for i, (label, conf_enum, fill, val_font) in enumerate(conf_items):
        row = conf_data_start + i
        count = sum(1 for f in recs if f.confidence == conf_enum)
        pct = (count / max(len(recs), 1)) * 100
        for col_idx, val in enumerate([label, count, f"{pct:.0f}%"], start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.fill = fill
            c.font = val_font
            c.alignment = Alignment(
                horizontal="left" if col_idx == 1 else "center",
                indent=1 if col_idx == 1 else 0,
                vertical="center",
            )
            c.border = _thin
        ws.row_dimensions[row].height = 16

    conf_data_end = conf_data_start + len(conf_items) - 1
    ws.row_dimensions[conf_data_end + 1].height = 10  # spacer

    # ── Top READY Decisions ───────────────────────────────────────────────
    top_hdr_row = conf_data_end + 2
    _section_header(top_hdr_row, "  TOP 10 READY DECISIONS", n_cols=5)
    _col_header(top_hdr_row + 1, ["Code", "Category", "Resource", "Confidence", "Rationale"])
    top_decisions = [f for f in recs if f.readiness == Readiness.READY][:10]
    for i, f in enumerate(top_decisions):
        row = top_hdr_row + 2 + i
        vm_display = f.vm_id.split("/")[-1] if "/" in f.vm_id else f.vm_id
        bg = _ALT_FILL if i % 2 == 0 else PatternFill()
        row_vals = [
            f.code,
            f.category.value.capitalize(),
            vm_display,
            f.confidence.value if f.confidence else "",
            f.rationale[:150] if f.rationale else "",
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = Font(size=9)
            cell.fill = bg
            cell.border = _thin
        if f.confidence:
            ws.cell(row=row, column=4).fill = _CONFIDENCE_FILL_MAP.get(f.confidence.value, bg)
        ws.row_dimensions[row].height = 16

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 2   # narrow gap before charts
    ws.column_dimensions["G"].width = 2

    # ── Charts ────────────────────────────────────────────────────────────
    # Anchored at column H — completely clear of all data in columns A-E.
    # Bar chart: Findings by Category & Readiness
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "Findings by Category & Readiness"
    bar.y_axis.title = "Count"
    bar.width = 20
    bar.height = 12
    bar.style = 10
    # Row 14 contains column headers (READY/LIKELY/DISCOVERY) → titles_from_data=True
    bar_data_ref = Reference(ws, min_col=2, max_col=4,
                              min_row=14, max_row=cat_data_end)
    bar_cats_ref = Reference(ws, min_col=1,
                              min_row=cat_data_start, max_row=cat_data_end)
    bar.add_data(bar_data_ref, titles_from_data=True)
    bar.set_categories(bar_cats_ref)
    ws.add_chart(bar, "H4")

    # Pie chart: Confidence distribution
    pie = PieChart()
    pie.title = "Recommendation Confidence"
    pie.width = 14
    pie.height = 11
    pie.style = 10
    pie_data_ref = Reference(ws, min_col=2,
                              min_row=conf_data_start, max_row=conf_data_end)
    pie_labs_ref = Reference(ws, min_col=1,
                              min_row=conf_data_start, max_row=conf_data_end)
    pie.add_data(pie_data_ref)
    pie.set_categories(pie_labs_ref)
    ws.add_chart(pie, "H27")


# ---------------------------------------------------------------------------
# Sheet 2: Decisions  (Phase B — RECOMMENDATION findings + Advisor)
# ---------------------------------------------------------------------------

_DECISIONS_COLS: list[tuple[str, int]] = [
    ("Code",                  16),
    ("Category",              14),
    ("Subcategory",           18),
    ("Status",                18),
    ("ResourceID",            52),
    ("Readiness",             14),
    ("Confidence",            14),
    ("Score",                  8),
    ("Evidence Sources",      30),
    ("Current",               22),
    ("Proposed",              22),
    ("Rationale",             50),
    ("Blockers to High",      36),
    ("Customer Inputs",       36),
]


def _sheet_decisions(
    wb: Workbook,
    findings: list[Finding],
    advisor: list[AdvisorRecommendation],
) -> None:
    ws = wb.create_sheet("Decisions")
    headers = [c[0] for c in _DECISIONS_COLS]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    recs = [f for f in findings if f.finding_type == FindingType.RECOMMENDATION]
    # Sort: READY first, then LIKELY, then by category
    readiness_order = {Readiness.READY: 0, Readiness.LIKELY: 1, Readiness.DISCOVERY: 2,
                       Readiness.INSUFFICIENT: 3}
    recs = sorted(recs, key=lambda f: (readiness_order.get(f.readiness, 9), f.category.value))

    for row_idx, f in enumerate(recs, start=2):
        alt = row_idx % 2 == 0
        vm_display = mask_subscription_ids_in_string(f.vm_id) if "/" in f.vm_id else f.vm_id
        row_vals = [
            f.code,
            f.category.value,
            f.subcategory.value if hasattr(f.subcategory, "value") else str(f.subcategory),
            "CSA to Review",
            vm_display,
            f.readiness.value,
            f.confidence.value if f.confidence else "",
            f.confidence_score if f.confidence_score is not None else "",
            " | ".join(f.evidence_sources),
            f.current or "",
            f.proposed or "",
            f.rationale or "",
            " | ".join(f.blockers_to_high),
            " | ".join(f.customer_inputs_needed),
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            # Colour-code Status col (4), Readiness col (6), Confidence col (7), Score col (8)
            if col_idx == 4:
                cell.fill = _CSA_FILL
            elif col_idx == 6:
                cell.fill = _READINESS_FILL_MAP.get(f.readiness.value, PatternFill())
            elif col_idx == 7 and f.confidence:
                cell.fill = _CONFIDENCE_FILL_MAP.get(f.confidence.value, PatternFill())
            elif col_idx == 8 and isinstance(val, int):
                # Tint the score cell with the same colour as the confidence band
                if f.confidence:
                    cell.fill = _CONFIDENCE_FILL_MAP.get(f.confidence.value, PatternFill())
                cell.font = Font(size=9, bold=True)
            elif alt:
                cell.fill = _ALT_FILL

    decisions_data_end = len(recs) + 1
    if decisions_data_end >= 1:
        _add_table(ws, "TblDecisions", decisions_data_end)

    # Advisor section (appended below a separator row)
    if advisor:
        sep_row = decisions_data_end + 3
        ws.cell(row=sep_row, column=1, value="Advisor SKU Recommendations").font = Font(
            bold=True, size=10, color="1F4E79")
        adv_hdr_row = sep_row + 1
        for idx, hdr in enumerate(
            ["Impact", "Description", "Current SKU", "Recommended SKU",
             "Subscription", "Resource ID"],
            start=1,
        ):
            c = ws.cell(row=adv_hdr_row, column=idx, value=hdr)
            c.fill = _HDR_FILL
            c.font = _HDR_FONT
        for i, rec in enumerate(advisor):
            row = adv_hdr_row + 1 + i
            alt = i % 2 == 0
            for col_idx, val in enumerate([
                getattr(rec, "impact", "") or "",
                getattr(rec, "short_description", "") or "",
                getattr(rec, "current_sku", "") or "",
                getattr(rec, "target_sku", "") or "",
                getattr(rec, "subscription_name", "") or "",
                mask_subscription_ids_in_string(getattr(rec, "resource_id", "") or ""),
            ], start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = Font(size=9)
                if alt:
                    cell.fill = _ALT_FILL

    for col_idx, (_, width) in enumerate(_DECISIONS_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# Sheet 3: Discovery Candidates  (Phase B)
# ---------------------------------------------------------------------------

_CANDIDATE_COLS: list[tuple[str, int]] = [
    ("Code",                  16),
    ("Category",              14),
    ("Subcategory",           18),
    ("ResourceID",            52),
    ("Rationale",             60),
    ("Customer Inputs Needed", 50),
]


def _sheet_discovery_candidates(wb: Workbook, findings: list[Finding]) -> None:
    ws = wb.create_sheet("Discovery Candidates")
    headers = [c[0] for c in _CANDIDATE_COLS]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    candidates = [f for f in findings if f.finding_type == FindingType.CANDIDATE]
    candidates = sorted(candidates, key=lambda f: f.category.value)

    for row_idx, f in enumerate(candidates, start=2):
        alt = row_idx % 2 == 0
        vm_display = mask_subscription_ids_in_string(f.vm_id) if "/" in f.vm_id else f.vm_id
        row_vals = [
            f.code,
            f.category.value,
            f.subcategory.value if hasattr(f.subcategory, "value") else str(f.subcategory),
            vm_display,
            f.rationale or "",
            " | ".join(f.customer_inputs_needed),
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if alt:
                cell.fill = _ALT_FILL

    _add_table(ws, "TblCandidates", len(candidates))

    for col_idx, (_, width) in enumerate(_CANDIDATE_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# Sheet 4: Evidence  (Phase B — replaces Performance Summary)
# ---------------------------------------------------------------------------

_EVIDENCE_METRICS = [
    ("Avg CPU %",         "Percentage CPU",             "avg"),
    ("P95 CPU %",         "Percentage CPU",             "p95"),
    ("Avg Mem Avail (GB)", "Available Memory Bytes",    "avg"),
    ("Disk Read IOps",    "Disk Read Operations/Sec",   "avg"),
    ("Disk Write IOps",   "Disk Write Operations/Sec",  "avg"),
]


def _sheet_evidence(
    wb: Workbook,
    vms: list[VmInventory],
    metrics_by_vm: dict[str, dict[str, VmMetrics]],
    findings: list[Finding],
) -> None:
    ws = wb.create_sheet("Evidence")

    # Build a lookup: resource_id → list[Finding code]
    vm_findings: dict[str, list[str]] = {}
    vm_evidence: dict[str, set[str]] = {}
    vm_confidence: dict[str, str] = {}
    for f in findings:
        vm_findings.setdefault(f.vm_id, []).append(f.code)
        vm_evidence.setdefault(f.vm_id, set()).update(f.evidence_sources)
        # Highest confidence per VM (HIGH > MEDIUM > LOW)
        conf_order = {"HIGH": 3, "MEDIUM": 2, "LOW": 1}
        if f.confidence:
            existing = vm_confidence.get(f.vm_id, "")
            if conf_order.get(f.confidence.value, 0) > conf_order.get(existing, 0):
                vm_confidence[f.vm_id] = f.confidence.value

    static_hdrs = ["VM Name", "Subscription", "Resource Group", "VM SKU", "vCPUs", "Mem (GB)"]
    metric_hdrs = [m[0] for m in _EVIDENCE_METRICS]
    extra_hdrs = ["Finding Codes", "Evidence Sources", "Best Confidence"]
    headers = static_hdrs + metric_hdrs + extra_hdrs
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_idx, vm in enumerate(vms, start=2):
        alt = row_idx % 2 == 0
        vm_met = metrics_by_vm.get(vm.resource_id, {})
        row_data: list[Any] = [
            vm.vm_name, vm.subscription_name, vm.resource_group,
            vm.vm_sku, vm.vcpus, vm.memory_gb,
        ]
        for _, metric_name, stat in _EVIDENCE_METRICS:
            m = vm_met.get(metric_name)
            val: float | None = getattr(m, stat, None) if m else None
            if val is not None and metric_name == "Available Memory Bytes":
                val = round(val / (1024 ** 3), 2)
            row_data.append(val)

        codes_str = " | ".join(vm_findings.get(vm.resource_id, []))
        sources_str = " | ".join(sorted(vm_evidence.get(vm.resource_id, set())))
        conf_str = vm_confidence.get(vm.resource_id, "")
        row_data.extend([codes_str, sources_str, conf_str])

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            if alt:
                cell.fill = _ALT_FILL
            # Colour CPU columns (positions 7-8 in 1-indexed)
            if col_idx == 7 and isinstance(value, float):
                _colour_util(cell, value)

    _add_table(ws, "TblEvidence", len(vms))

    widths = [28, 26, 24, 20, 8, 10, 12, 12, 16, 12, 12, 36, 36, 16]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# Sheet 8: Source Coverage  (Phase B)
# ---------------------------------------------------------------------------

def _sheet_source_coverage(
    wb: Workbook,
    vms: list[VmInventory],
    enriched_metrics: list[EnrichedVmMetrics],
) -> None:
    ws = wb.create_sheet("Source Coverage")

    # Build lookup: vm_name (lower) → EnrichedVmMetrics
    enriched_by_name: dict[str, EnrichedVmMetrics] = {
        e.vm_name.lower(): e for e in enriched_metrics
    }

    headers = [
        "VM Name", "VM SKU", "Subscription", "Resource Group",
        "Source Tool", "Confidence Tier", "Has OS Data",
        "Has JVM Data", "Has .NET Data", "Has SQL Data", "Data Points",
    ]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_idx, vm in enumerate(vms, start=2):
        alt = row_idx % 2 == 0
        enriched = enriched_by_name.get(vm.vm_name.lower())
        if enriched:
            source_tool = enriched.source_tool
            tier = enriched.confidence_tier
            has_os = "Yes" if enriched.has_os_data else "No"
            has_jvm = "Yes" if enriched.has_jvm_data else "No"
            has_dotnet = "Yes" if enriched.has_dotnet_data else "No"
            has_sql = "Yes" if enriched.has_sql_data else "No"
            data_pts = len(enriched.data_points)
        else:
            source_tool = "Platform only"
            tier = "platform-only"
            has_os = has_jvm = has_dotnet = has_sql = "No"
            data_pts = 0

        row_vals = [
            vm.vm_name, vm.vm_sku, vm.subscription_name, vm.resource_group,
            source_tool, tier, has_os, has_jvm, has_dotnet, has_sql, data_pts,
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            if alt:
                cell.fill = _ALT_FILL
            # Colour-code Confidence Tier (col 6)
            if col_idx == 6:
                if tier == "workload-aware":
                    cell.fill = _HIGH_FILL
                elif tier == "os-aware":
                    cell.fill = _MED_FILL
                else:
                    cell.fill = _LOW_FILL

    _add_table(ws, "TblSourceCoverage", len(vms))

    for col_idx, width in enumerate([28, 20, 26, 24, 18, 18, 12, 12, 12, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# Sheet 9: Anomalies  (Phase B — platform vs external disagreements)
# ---------------------------------------------------------------------------

_ANOMALY_THRESHOLD_PCT: float = 10.0


def _sheet_anomalies(
    wb: Workbook,
    vms: list[VmInventory],
    metrics_by_vm: dict[str, dict[str, VmMetrics]],
    enriched_metrics: list[EnrichedVmMetrics],
) -> None:
    ws = wb.create_sheet("Anomalies")

    headers = [
        "VM Name", "VM SKU", "Subscription", "Resource Group",
        "Metric", "Platform Avg", "External Avg", "Delta (pts)", "Interpretation",
    ]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    enriched_by_name: dict[str, EnrichedVmMetrics] = {
        e.vm_name.lower(): e for e in enriched_metrics
    }

    # Compare pairs: platform CPU % vs os.cpu.percent, mem available vs os.memory.used_percent
    _comparisons: list[tuple[str, str, str, str]] = [
        # (display_name, platform_metric, ext_metric, direction)
        ("CPU %", "Percentage CPU", "os.cpu.percent", "both"),
    ]

    row_idx = 2
    for vm in vms:
        enriched = enriched_by_name.get(vm.vm_name.lower())
        if not enriched:
            continue
        vm_met = metrics_by_vm.get(vm.resource_id, {})
        ext_lookup: dict[str, Any] = {dp.metric_name: dp for dp in enriched.data_points}

        for display, plat_name, ext_name, _ in _comparisons:
            plat_m = vm_met.get(plat_name)
            ext_dp = ext_lookup.get(ext_name)
            if plat_m is None or ext_dp is None:
                continue
            plat_val = plat_m.avg
            ext_val = ext_dp.avg_value
            if plat_val is None or ext_val is None:
                continue
            delta = abs(plat_val - ext_val)
            if delta < _ANOMALY_THRESHOLD_PCT:
                continue
            direction = (
                "Platform higher — check agent coverage"
                if plat_val > ext_val
                else "External higher — check platform metric gap"
            )
            alt = row_idx % 2 == 0
            for col_idx, val in enumerate([
                vm.vm_name, vm.vm_sku, vm.subscription_name, vm.resource_group,
                display,
                round(plat_val, 2), round(ext_val, 2), round(delta, 2),
                direction,
            ], start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = _THIN_BORDER
                cell.font = Font(size=9)
                if alt:
                    cell.fill = _ALT_FILL
                if col_idx == 8:
                    cell.fill = _RED_FILL if delta >= 20 else _YELLOW_FILL
            row_idx += 1

    data_rows = row_idx - 2
    if data_rows > 0:
        _add_table(ws, "TblAnomalies", data_rows)

    for col_idx, width in enumerate([28, 20, 26, 24, 14, 14, 14, 14, 40], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# Sheet 10: Data Gap Register  (Phase B — VMs with no external enrichment)
# ---------------------------------------------------------------------------

def _sheet_data_gap_register(
    wb: Workbook,
    vms: list[VmInventory],
    enriched_metrics: list[EnrichedVmMetrics],
) -> None:
    ws = wb.create_sheet("Data Gap Register")

    headers = [
        "VM Name", "VM SKU", "Subscription", "Resource Group", "Region",
        "Gap Reason", "Uplift Action",
    ]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    enriched_names: set[str] = {e.vm_name.lower() for e in enriched_metrics}
    os_aware_names: set[str] = {
        e.vm_name.lower() for e in enriched_metrics if e.has_os_data
    }

    row_idx = 2
    for vm in vms:
        vn = vm.vm_name.lower()
        if vn in os_aware_names:
            continue  # has OS data — not a gap (may still be MEDIUM if no workload metrics)
        alt = row_idx % 2 == 0
        if vn not in enriched_names:
            gap_reason = "No external monitoring data — platform metrics only"
            uplift = "Install monitoring agent + export canonical CSV"
        else:
            gap_reason = "External data present but no OS-level metrics"
            uplift = "Verify agent has os.* metric collection enabled"

        for col_idx, val in enumerate([
            vm.vm_name, vm.vm_sku, vm.subscription_name, vm.resource_group, vm.region,
            gap_reason, uplift,
        ], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            if alt:
                cell.fill = _ALT_FILL
        row_idx += 1

    data_rows = row_idx - 2
    if data_rows > 0:
        _add_table(ws, "TblDataGap", data_rows)

    for col_idx, width in enumerate([28, 20, 26, 24, 16, 48, 48], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# Sheet 11: Fleet Inventory  (renamed from VM Inventory)

_INVENTORY_COLS: list[tuple[str, str, int]] = [
    # (header, field_name_or_special, width)
    ("VM Name", "vm_name", 28),
    ("Subscription", "subscription_name", 30),
    ("Subscription ID", "masked_subscription_id", 42),
    ("Resource Group", "resource_group", 28),
    ("Region", "region", 16),
    ("VM SKU", "vm_sku", 22),
    ("vCPUs", "vcpus", 8),
    ("Memory (GB)", "memory_gb", 12),
    ("OS Type", "os_type", 12),
    ("OS Version", "os_version", 20),
    ("Power State", "power_state", 18),
    ("Image Publisher", "image_publisher", 26),
    ("Image Offer", "image_offer", 22),
    ("Image SKU", "image_sku", 24),
    ("Image Version", "image_version", 22),
    ("Avail. Zone", "availability_zone", 12),
    ("NIC Count", "nic_count", 10),
    ("Disk Count", "disk_count", 10),
    ("Disk Sizes (GB)", "disk_sizes_gb", 22),
    ("VMSS Name", "vmss_name", 24),
    ("Availability Set", "availability_set_name", 24),
    ("Resource ID", "masked_resource_id", 80),
    # Analyst-editable (light blue)
    ("Workload", "workload", 18),
    ("Application", "application", 18),
    ("Environment", "environment", 16),
    ("Criticality", "criticality", 14),
    ("Owner", "owner", 20),
    ("Custom", "custom", 20),
]

_CSA_START_COL = 23  # first analyst-editable column (1-indexed); shifts when columns added before it


def _sheet_inventory(wb: Workbook, vms: list[VmInventory]) -> None:
    ws = wb.create_sheet("Fleet Inventory")

    headers = [c[0] for c in _INVENTORY_COLS]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_idx, vm in enumerate(vms, start=2):
        alt = row_idx % 2 == 0
        for col_idx, (_, field, _) in enumerate(_INVENTORY_COLS, start=1):
            val: Any
            if field == "masked_subscription_id":
                val = vm.masked_subscription_id()
            elif field == "masked_resource_id":
                val = vm.masked_resource_id()
            elif field == "disk_sizes_gb":
                val = ", ".join(str(int(s)) for s in vm.disk_sizes_gb)
            else:
                val = getattr(vm, field, None)
                if val is None:
                    val = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=False)

            is_csa = col_idx >= _CSA_START_COL
            if is_csa:
                cell.fill = _CSA_FILL
            elif alt:
                cell.fill = _ALT_FILL

    for col_idx, (_, _, width) in enumerate(_INVENTORY_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    _add_table(ws, "TblInventory", len(vms))


# ---------------------------------------------------------------------------
# Sheet 2: Performance Summary
# ---------------------------------------------------------------------------

_PERF_METRICS = [
    ("Avg CPU %",       "Percentage CPU", "avg"),
    ("P95 CPU %",       "Percentage CPU", "p95"),
    ("P99 CPU %",       "Percentage CPU", "p99"),
    ("Max CPU % (Peak)", "Percentage CPU", "max"),
    ("Min CPU %",       "Percentage CPU", "min"),
    ("Avg Mem Avail (GB)", "Available Memory Bytes", "avg"),
    ("Disk Read IOps", "Disk Read Operations/Sec", "avg"),
    ("Disk Write IOps", "Disk Write Operations/Sec", "avg"),
    ("Net In (GB)", "Network In Total", "avg"),
    ("Net Out (GB)", "Network Out Total", "avg"),
]


def _sheet_perf_summary(
    wb: Workbook,
    vms: list[VmInventory],
    metrics_by_vm: dict[str, dict[str, VmMetrics]],
) -> None:
    ws = wb.create_sheet("Performance Summary")

    static_headers = ["VM Name", "Subscription", "Resource Group", "VM SKU", "vCPUs", "Mem (GB)"]
    metric_headers = [m[0] for m in _PERF_METRICS]
    headers = static_headers + metric_headers
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_idx, vm in enumerate(vms, start=2):
        alt = row_idx % 2 == 0
        vm_met = metrics_by_vm.get(vm.resource_id, {})
        row_data: list[Any] = [
            vm.vm_name, vm.subscription_name, vm.resource_group,
            vm.vm_sku, vm.vcpus, vm.memory_gb,
        ]
        for _, metric_name, stat in _PERF_METRICS:
            m = vm_met.get(metric_name)
            val: float | None = getattr(m, stat, None) if m else None
            if val is not None and metric_name == "Available Memory Bytes":
                val = round(val / (1024 ** 3), 2)
            elif val is not None and metric_name in ("Network In Total", "Network Out Total"):
                val = round(val / (1024 ** 3), 4)
            row_data.append(val)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            if alt:
                cell.fill = _ALT_FILL

            # Format and colour-code CPU % columns (Avg, P95, P99, Max, Min)
            if col_idx in (7, 8, 9, 10, 11) and value is not None:
                cell.number_format = "0.00"
                if value < 40:
                    cell.fill = _GREEN_FILL
                elif value < 70:
                    cell.fill = _YELLOW_FILL
                else:
                    cell.fill = _RED_FILL

    _auto_width(ws, headers)
    _add_table(ws, "TblPerfSummary", len(vms))


# ---------------------------------------------------------------------------
# Sheet 3–6: SKU flat (normalized row-per-group+sku) sheets
# ---------------------------------------------------------------------------

def _sheet_sku_flat(
    wb: Workbook,
    vms: list[VmInventory],
    metrics_by_vm: dict[str, dict[str, VmMetrics]],
    sheet_name: str,
    group_model_fields: list[str],
    col_headers: list[str],
    col_widths: list[int],
    table_name: str,
) -> None:
    """Flat normalized view: one row per (group_fields..., vm_sku) combination.

    Columns: <group_fields...> | VM SKU | VM Count | Avg CPU % | Avg Mem %
    """
    ws = wb.create_sheet(sheet_name)
    headers = col_headers + ["VM SKU", "VM Count", "Avg CPU %", "P95 CPU %", "P99 CPU %", "Max CPU % (Peak)", "Min CPU %", "Avg Mem %"]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Group vms by (group_fields..., vm_sku)
    row_groups: dict[tuple, list[VmInventory]] = {}
    for vm in vms:
        key = tuple(getattr(vm, f, None) or "(unknown)" for f in group_model_fields) + (vm.vm_sku,)
        row_groups.setdefault(key, []).append(vm)

    n_group = len(group_model_fields)
    avg_cpu_col = n_group + 3  # group cols + VM SKU + VM Count → Avg CPU
    mem_col = avg_cpu_col + 5  # after Avg, P95, P99, Max, Min

    for row_idx, (key, group_vms) in enumerate(sorted(row_groups.items()), start=2):
        alt = row_idx % 2 == 0
        *group_vals, sku = key
        avg_cpu = _avg_metric(group_vms, metrics_by_vm, "Percentage CPU", "avg")
        p95_cpu = _avg_metric(group_vms, metrics_by_vm, "Percentage CPU", "p95")
        p99_cpu = _avg_metric(group_vms, metrics_by_vm, "Percentage CPU", "p99")
        max_cpu = _avg_metric(group_vms, metrics_by_vm, "Percentage CPU", "max")
        min_cpu = _avg_metric(group_vms, metrics_by_vm, "Percentage CPU", "min")
        avg_mem = _avg_mem_pct(group_vms, metrics_by_vm)
        row_data = list(group_vals) + [sku, len(group_vms), avg_cpu, p95_cpu, p99_cpu, max_cpu, min_cpu, avg_mem]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            if alt:
                cell.fill = _ALT_FILL
        for cpu_col in range(avg_cpu_col, avg_cpu_col + 5):
            _colour_util(ws.cell(row=row_idx, column=cpu_col), row_data[cpu_col - 1])
        _colour_util(ws.cell(row=row_idx, column=mem_col), avg_mem)

    all_widths = col_widths + [22, 10, 13, 13, 13, 17, 13, 13]
    for col_idx, w in enumerate(all_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    if row_groups:
        _add_table(ws, table_name, len(row_groups))


# ---------------------------------------------------------------------------
# Recommendation action helper
# ---------------------------------------------------------------------------

def _rec_action(rec: "VmRecommendation") -> str:
    """Plain-English recommended action to present to the Workload Owner/SMEs."""
    if rec.category == "underutilized":
        if rec.recommended_sku:
            return f"Resize to {rec.recommended_sku} or decommission if unused"
        return "Review usage — decommission or resize to smallest available SKU"
    if rec.category == "right-size":
        if rec.recommended_sku:
            return f"Resize {rec.current_sku} → {rec.recommended_sku}"
        return "Resize to a smaller SKU (no smaller SKU found in current region)"
    if rec.category == "PaaS-candidate":
        return "Migrate to Azure App Service, Container Apps, or Azure SQL"
    return ""


# ---------------------------------------------------------------------------
# Sheet 7: Optimizations
# ---------------------------------------------------------------------------

# Order matches the reporting spec exactly:
#   Priority, Recommendation, Category, Subcategory, Resource ID, Members,
#   Current SKU/Resource Type, Recommended SKU/Resource Type,
#   Reason, Estimated Optimization, Override, Notes, Confidence, Evidence
_REC_COLS = [
    ("Priority", 12),
    ("Recommendation", 42),
    ("Category", 26),
    ("Subcategory", 22),
    ("Resource ID", 60),
    ("Members", 10),
    ("Current SKU / Resource Type", 32),
    ("Recommended SKU / Resource Type", 32),
    ("Reason", 50),
    ("Estimated Optimization", 28),
    ("Override", 14),
    ("Notes", 24),
    ("Confidence", 18),
    ("Evidence", 52),
]

_PRIORITY_FILLS = {
    "critical": PatternFill("solid", fgColor="C00000"),  # red
    "high":     PatternFill("solid", fgColor="ED7D31"),  # orange
    "medium":   PatternFill("solid", fgColor="FFC000"),  # yellow / gold
    "low":      PatternFill("solid", fgColor="70AD47"),  # green
}
_PRIORITY_FONTS = {
    "critical": Font(color="FFFFFF", bold=True, size=9),
    "high":     Font(color="FFFFFF", bold=True, size=9),
    "medium":   Font(color="000000", bold=True, size=9),
    "low":      Font(color="FFFFFF", bold=True, size=9),
}


def _sheet_recommendations(wb: Workbook, recommendations: list[VmRecommendation]) -> None:
    ws = wb.create_sheet("Optimizations")
    headers = [c[0] for c in _REC_COLS]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Data validation for Override column (column 11).
    # sqref is set AFTER the row loop so we have the final row count; the DV
    # is only registered with the sheet if there are rows to cover.
    dv = DataValidation(
        type="list",
        formula1='"accept,reject,defer"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error="Choose: accept, reject, or defer",
    )

    for row_idx, rec in enumerate(recommendations, start=2):
        alt = row_idx % 2 == 0
        current = rec.current_sku or rec.current_resource_type or ""
        recommended = rec.recommended_sku or rec.recommended_resource_type or ""
        # When the rec covers multiple VMs (VMSS, AVSet, …) show the parent
        # resource ID and the member count; standalone VMs show the VM ID and 1.
        if rec.member_count > 1 and rec.parent_resource_id:
            displayed_id = rec.masked_parent_resource_id() or rec.masked_resource_id()
        else:
            displayed_id = rec.masked_resource_id()
        row_data = [
            rec.priority,
            rec.recommendation,
            rec.category,
            rec.subcategory,
            displayed_id,
            rec.member_count,
            current,
            recommended,
            rec.reason,
            rec.estimated_optimization,
            rec.manual_override or "",
            rec.notes or "",
            rec.confidence,
            " | ".join(rec.evidence) if rec.evidence else "",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=col_idx in (2, 9, 10, 12, 14),
            )
            if alt:
                cell.fill = _ALT_FILL

        # Priority cell — bold colour-coded badge
        prio_cell = ws.cell(row=row_idx, column=1)
        fill = _PRIORITY_FILLS.get((rec.priority or "").lower())
        font = _PRIORITY_FONTS.get((rec.priority or "").lower())
        if fill and font:
            prio_cell.fill = fill
            prio_cell.font = font
            prio_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Override + Notes are analyst-editable (columns 11, 12)
        for col_idx in (11, 12):
            ws.cell(row=row_idx, column=col_idx).fill = _CSA_FILL

        # Confidence column (13) — colour-coded by tier
        conf_cell = ws.cell(row=row_idx, column=13)
        if rec.confidence == "workload-aware":
            conf_cell.fill = _GREEN_FILL
        elif rec.confidence == "os-aware":
            conf_cell.fill = _YELLOW_FILL

    # Register the DataValidation only if there are rows; set sqref as a
    # compact Excel range (e.g. "K2:K101") — NOT per-cell space-separated
    # list, which grows to thousands of chars and causes "Catastrophic failure"
    # in Excel's XML parser.
    if recommendations:
        dv.sqref = f"K2:K{len(recommendations) + 1}"
        ws.add_data_validation(dv)

    for col_idx, (_, width) in enumerate(_REC_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    _add_table(ws, "TblRecommendations", len(recommendations))


# ---------------------------------------------------------------------------
# Sheet 2: Quota Utilization
# ---------------------------------------------------------------------------

_QUOTA_COLS = [
    ("Subscription", 30),
    ("Region", 16),
    ("Resource Type", 30),
    ("Display Name", 35),
    ("Current Usage", 14),
    ("Quota Limit", 14),
    ("Utilization %", 14),
    ("30d Peak %", 12),
    ("Alloc. Failures (30d)", 22),
    ("PAYG Default", 14),
    ("Raised?", 10),
    ("Alert", 8),
]


def _sheet_quota(wb: Workbook, quota_items: list[QuotaItem]) -> None:
    ws = wb.create_sheet("Quota Posture")
    headers = [c[0] for c in _QUOTA_COLS]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_idx, item in enumerate(quota_items, start=2):
        alt = row_idx % 2 == 0
        row_data = [
            item.subscription_name,
            item.region,
            item.resource_type,
            item.display_name,
            item.current_usage,
            item.quota_limit,
            item.utilization_pct,
            item.peak_usage_pct_30d,
            item.allocation_failures_30d if item.allocation_failures_30d else "",
            item.subscription_default if item.subscription_default is not None else "",
            "\u2713" if (
                item.subscription_default is not None
                and item.quota_limit > item.subscription_default
            ) else "",
            "YES" if item.alert else "",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="top")
            if item.alert:
                cell.fill = _RED_FILL
            elif item.allocation_failures_30d > 0 and col_idx == 9:
                cell.fill = _RED_FILL
            elif col_idx == 10 and isinstance(val, int) and val == 0:
                # Highlight zero PAYG default (e.g. restricted GPU family)
                cell.fill = _YELLOW_FILL
            elif alt:
                cell.fill = _ALT_FILL
        # Colour-code utilization % cell (col 7)
        util_cell = ws.cell(row=row_idx, column=7)
        if not item.alert:  # only recolour if not already red
            _colour_util(util_cell, item.utilization_pct)
        # Colour-code peak usage % cell (col 8)
        peak_cell = ws.cell(row=row_idx, column=8)
        if item.peak_usage_pct_30d is not None and not item.alert:
            _colour_util(peak_cell, item.peak_usage_pct_30d)

    for col_idx, (_, width) in enumerate(_QUOTA_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    if quota_items:
        _add_table(ws, "TblQuota", len(quota_items))


def _read_quota_sheet(wb) -> list[QuotaItem]:
    # Support Phase B name ("Quota Posture"), original ("Quota Utilization"), and old typo
    sheet_name = next(
        (n for n in ("Quota Posture", "Quota Utilization", "Quota Utilisation")
         if n in wb.sheetnames),
        None,
    )
    if sheet_name is None:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    # Detect column layout (10-col new vs 8-col legacy)
    header = [str(c or "").strip() for c in rows[0]]
    has_peak = "30d Peak %" in header
    has_failures = "Alloc. Failures (30d)" in header
    has_payg = "PAYG Default" in header
    items: list[QuotaItem] = []
    for row in rows[1:]:
        if not any(row):
            continue
        try:
            if has_payg:
                # New 12-column layout: PAYG Default at [9], Raised? at [10], Alert at [11]
                alert_val = str(row[11] or "").upper()
                peak = float(row[7]) if row[7] is not None else None
                failures = int(row[8] or 0)
                payg_default = int(row[9]) if row[9] is not None and row[9] != "" else None
            elif has_peak and has_failures:
                # 10-column layout
                alert_val = str(row[9] or "").upper()
                peak = float(row[7]) if row[7] is not None else None
                failures = int(row[8] or 0)
                payg_default = None
            else:
                # Legacy 8-column layout
                alert_val = str(row[7] or "").upper()
                peak = None
                failures = 0
                payg_default = None
            items.append(
                QuotaItem(
                    subscription_id="",
                    subscription_name=str(row[0] or ""),
                    region=str(row[1] or ""),
                    resource_type=str(row[2] or ""),
                    display_name=str(row[3] or ""),
                    current_usage=int(row[4] or 0),
                    quota_limit=int(row[5] or 0),
                    utilization_pct=float(row[6] or 0),
                    alert=(alert_val == "YES"),
                    peak_usage_pct_30d=peak,
                    allocation_failures_30d=failures,
                    subscription_default=payg_default,
                )
            )
        except Exception:
            continue
    return items


# ---------------------------------------------------------------------------
# Sheet 9: Raw Metrics
# ---------------------------------------------------------------------------

def _sheet_raw_metrics(wb: Workbook, metrics: list[VmMetrics]) -> None:
    ws = wb.create_sheet("Raw Metrics")
    headers = ["Resource ID", "Metric", "Avg", "P50", "P95", "P99", "Max", "Min", "Data Points"]
    _write_header(ws, headers)

    from cloudopt.models import mask_subscription_ids_in_string

    for row_idx, m in enumerate(metrics, start=2):
        row = [
            mask_subscription_ids_in_string(m.resource_id),
            m.metric_name,
            _fmt(m.avg), _fmt(m.p50), _fmt(m.p95), _fmt(m.p99), _fmt(m.max), _fmt(m.min),
            len(m.time_series),
        ]
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(size=9)

    _auto_width(ws, headers)


# ---------------------------------------------------------------------------
# Sheet 10: Application Insights
# ---------------------------------------------------------------------------

def _sheet_appinsights(
    wb: Workbook,
    components: list[AppInsightsInventory],
    metrics: list[AppInsightsMetrics],
) -> None:
    """App Insights inventory + summarised metrics in a single sheet."""

    ws = wb.create_sheet("App Insights")

    # Build a lookup: resource_id → {metric_name: AppInsightsMetrics}
    metrics_lookup: dict[str, dict[str, AppInsightsMetrics]] = {}
    for m in metrics:
        metrics_lookup.setdefault(m.resource_id, {})[m.metric_name] = m

    # Ordered metric columns to show in the sheet
    METRIC_COLS: list[tuple[str, str]] = [
        ("availabilityResults/availabilityPercentage", "Avail %"),
        ("requests/count",                             "Req Count"),
        ("requests/duration",                          "Req Dur ms"),
        ("requests/failed",                            "Failed Req"),
        ("exceptions/count",                           "Exceptions"),
        ("performanceCounters/processCpuPercentage",   "Proc CPU %"),
        ("performanceCounters/processPrivateBytes",    "Priv Bytes"),
        ("performanceCounters/memoryAvailableBytes",   "Avail Mem"),
        ("jvm/memory/heap/used",                       "JVM Heap Used"),
        ("jvm/memory/heap/max",                        "JVM Heap Max"),
        ("jvm/gc/pause",                               "JVM GC ms"),
        ("jvm/gc/count",                               "JVM GC Cnt"),
        ("jvm/threads/count",                          "JVM Threads"),
    ]

    headers = [
        "Component", "Subscription", "Resource Group", "Region",
        "Kind", "Type", "Workspace?",
    ] + [f"{label} (avg)" for _, label in METRIC_COLS]

    _write_header(ws, headers)
    col_widths = [32, 26, 22, 16, 12, 10, 12] + [14] * len(METRIC_COLS)
    for col_i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = width

    for row_i, comp in enumerate(components, start=2):
        comp_metrics = metrics_lookup.get(comp.resource_id, {})
        metric_avgs = [
            comp_metrics[m_name].avg if m_name in comp_metrics else None
            for m_name, _ in METRIC_COLS
        ]
        row_data = [
            comp.component_name,
            comp.subscription_name,
            comp.resource_group,
            comp.region,
            comp.kind,
            comp.application_type,
            "Yes" if comp.workspace_resource_id else "No",
        ] + metric_avgs

        fill = _ALT_FILL if row_i % 2 == 0 else None
        for col_i, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=value)
            cell.border = _THIN_BORDER
            if fill:
                cell.fill = fill
            if isinstance(value, float) and col_i > 7:
                cell.number_format = "#,##0.00"

        # Pre-compute failed request ratio for availability highlighting
        _req_count_m  = comp_metrics.get("requests/count")
        _req_failed_m = comp_metrics.get("requests/failed")
        _failed_ratio: float | None = None
        if (
            _req_count_m and _req_failed_m
            and _req_count_m.avg and _req_count_m.avg > 0
        ):
            _failed_ratio = (_req_failed_m.avg or 0.0) / _req_count_m.avg

        # Apply per-metric highlighting (overrides alt-row fill)
        _metric_col_start = 8  # columns 1-7 are static fields
        for _met_i, (_m_name, _) in enumerate(METRIC_COLS):
            _col_i = _metric_col_start + _met_i
            _val   = metric_avgs[_met_i]
            if _val is None:
                continue
            _hl: Any = None
            if _m_name == "availabilityResults/availabilityPercentage":
                if _val < 95:
                    _hl = _RED_FILL
                elif _val < 99:
                    _hl = _YELLOW_FILL
            elif _m_name == "requests/duration":
                if _val > 5000:
                    _hl = _RED_FILL
                elif _val > 2000:
                    _hl = _YELLOW_FILL
            elif _m_name == "requests/failed" and _failed_ratio is not None:
                if _failed_ratio > 0.05:
                    _hl = _RED_FILL
                elif _failed_ratio > 0.01:
                    _hl = _YELLOW_FILL
            elif _m_name == "exceptions/count" and _val > 0:
                _hl = _YELLOW_FILL
            if _hl is not None:
                ws.cell(row=row_i, column=_col_i).fill = _hl

    if components:
        tbl_ref = f"A1:{get_column_letter(len(headers))}{len(components) + 1}"
        tbl = Table(displayName="TblAppInsights", ref=tbl_ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True
        )
        ws.add_table(tbl)


# ---------------------------------------------------------------------------
# Sheet 0: Workload Information (filled in with Workload Owner/SMEs)
# ---------------------------------------------------------------------------

_WORKLOAD_ROWS: list[tuple[str, str]] = [
    ("Workload Name",                  "workload_name"),
    ("Azure Cloud",                    "azure_cloud"),
    ("Primary Region",                 "primary_region"),
    ("Secondary/DR Region",            "secondary_dr_region"),
    ("Criticality for the Business",   "business_criticality"),
    ("Availability/DR Design Pattern", "availability_dr_pattern"),
    ("SLA",                            "sla"),
    ("RPO",                            "rpo"),
    ("RTO",                            "rto"),
]


def _sheet_workload_info(wb: Workbook, info: WorkloadInfo) -> None:
    """Two-column table for capturing workload context from the Workload Owner/SMEs."""
    ws = wb.create_sheet("Workload Information")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 60
    ws.freeze_panes = "A2"

    headers = ["Property", "Workload Information"]
    _write_header(ws, headers)

    for row_idx, (label, field) in enumerate(_WORKLOAD_ROWS, start=2):
        prop_cell = ws.cell(row=row_idx, column=1, value=label)
        prop_cell.font = Font(bold=True, size=10)
        prop_cell.border = _THIN_BORDER
        prop_cell.alignment = Alignment(vertical="center", wrap_text=True)

        val_cell = ws.cell(row=row_idx, column=2, value=getattr(info, field, "") or "")
        val_cell.border = _THIN_BORDER
        val_cell.fill = _CSA_FILL
        val_cell.alignment = Alignment(vertical="center", wrap_text=True)
        val_cell.font = Font(size=10)
        ws.row_dimensions[row_idx].height = 22


# ---------------------------------------------------------------------------
# Sheet: Advisor SKU-change Recommendations
# ---------------------------------------------------------------------------

_ADVISOR_COLS: list[tuple[str, int]] = [
    ("Subscription",         28),
    ("Resource Group",       24),
    ("Resource Name",        28),
    ("Resource Type",        38),
    ("Category",             14),
    ("Impact",               10),
    ("Recommendation",       60),
    ("Current SKU",          18),
    ("Recommended SKU",      18),
    ("Annual Savings (USD)", 18),
    ("Last Updated",         22),
    ("Resource ID",          70),
]


def _sheet_advisor(wb: Workbook, recs: list[AdvisorRecommendation]) -> None:
    """Azure Advisor recommendations that suggest a SKU change."""
    ws = wb.create_sheet("Advisor SKU Changes")
    headers = [c[0] for c in _ADVISOR_COLS]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for col_idx, (_, width) in enumerate(_ADVISOR_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, rec in enumerate(recs, start=2):
        alt = row_idx % 2 == 0
        row_data = [
            rec.subscription_name,
            rec.resource_group,
            rec.impacted_resource_name,
            rec.impacted_resource_type,
            rec.category,
            rec.impact,
            rec.short_description,
            rec.current_sku,
            rec.recommended_sku,
            rec.annual_savings_usd,
            rec.last_updated,
            rec.masked_impacted_resource_id(),
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=col_idx in (7, 12))
            if alt:
                cell.fill = _ALT_FILL
        impact_cell = ws.cell(row=row_idx, column=6)
        impact_val = (rec.impact or "").lower()
        if impact_val == "high":
            impact_cell.fill = _RED_FILL
        elif impact_val == "medium":
            impact_cell.fill = _YELLOW_FILL
        elif impact_val == "low":
            impact_cell.fill = _GREEN_FILL

    if recs:
        _add_table(ws, "TblAdvisor", len(recs))


# Sheet 11: Subscriptions Zone Mapping
# ---------------------------------------------------------------------------

_ZONE_MAPPING_COLS: list[tuple[str, int]] = [
    ("Tenant ID", 36),
    ("Subscription ID", 42),
    ("Subscription Name", 36),
    ("Location", 20),
    ("Logical Zone", 14),
    ("Physical Zone", 24),
    ("Physical Zone Name", 26),
]


def _sheet_zone_mapping(
    wb: Workbook, zone_mappings: list[SubscriptionZoneMapping]
) -> None:
    """Physical-to-logical AZ mapping per subscription and location."""
    ws = wb.create_sheet("SubscriptionsZoneMapping")
    headers = [c[0] for c in _ZONE_MAPPING_COLS]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for col_idx, (_, width) in enumerate(_ZONE_MAPPING_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, zm in enumerate(zone_mappings, start=2):
        alt = row_idx % 2 == 0
        row_data = [
            zm.tenant_id,
            mask_subscription_id(zm.subscription_id),
            zm.subscription_name,
            zm.location,
            zm.logical_zone,
            zm.physical_zone,
            zm.physical_zone_name,
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="top")
            if alt:
                cell.fill = _ALT_FILL

    if zone_mappings:
        _add_table(ws, "TblZoneMapping", len(zone_mappings))


# Sheet 13: Monitoring Data
# ---------------------------------------------------------------------------

_MONITORING_COLS = [
    ("VM Name", 28),
    ("Source Tool", 16),
    ("Metric Group", 16),
    ("Metric Name", 34),
    ("Avg Value", 12),
    ("P95 Value", 12),
    ("Max Value", 12),
    ("Unit", 14),
    ("Confidence", 18),
]


def _sheet_monitoring_data(
    wb: Workbook,
    enriched_metrics: list[EnrichedVmMetrics],
) -> None:
    """Write monitoring data in tall format — one row per metric per VM."""
    from cloudopt.enrichment.schema import METRIC_GROUPS

    ws = wb.create_sheet("Monitoring Data")
    headers = [c[0] for c in _MONITORING_COLS]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Build reverse mapping: metric_name → group
    metric_to_group: dict[str, str] = {}
    for group_name, metric_names in METRIC_GROUPS.items():
        for mn in metric_names:
            metric_to_group[mn] = group_name

    row_idx = 2
    for enriched in enriched_metrics:
        for dp in enriched.data_points:
            alt = row_idx % 2 == 0
            group_name = metric_to_group.get(dp.metric_name, "other")
            row_data = [
                enriched.vm_name,
                dp.source_tool,
                group_name,
                dp.metric_name,
                dp.avg_value,
                dp.p95_value,
                dp.max_value,
                dp.unit,
                enriched.confidence_tier,
            ]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = _THIN_BORDER
                cell.font = Font(size=9)
                cell.alignment = Alignment(vertical="top")
                if alt:
                    cell.fill = _ALT_FILL
            # Confidence column colour
            conf_cell = ws.cell(row=row_idx, column=9)
            if enriched.confidence_tier == "workload-aware":
                conf_cell.fill = _GREEN_FILL
            elif enriched.confidence_tier == "os-aware":
                conf_cell.fill = _YELLOW_FILL
            row_idx += 1

    for col_idx, (_, width) in enumerate(_MONITORING_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row_count = row_idx - 2
    if row_count > 0:
        _add_table(ws, "TblMonitoringData", row_count)


# ---------------------------------------------------------------------------
# Sheet 13: Inventory (full ARG resource list, tags excluded)
# ---------------------------------------------------------------------------

_RESOURCES_COLS: list[tuple[str, str, int]] = [
    # (header, field_name_or_special, width)
    ("Name", "name", 36),
    ("Type", "resource_type", 52),
    ("Resource Group", "resource_group", 28),
    ("Subscription", "subscription_name", 30),
    ("Subscription ID", "masked_subscription_id", 42),
    ("Region", "location", 16),
    ("Kind", "kind", 20),
    ("SKU Name", "sku_name", 22),
    ("SKU Tier", "sku_tier", 16),
    ("Plan Name", "plan_name", 22),
    ("Plan Publisher", "plan_publisher", 22),
    ("Plan Product", "plan_product", 22),
    ("Zones", "zones", 12),
    ("Managed By", "managed_by", 50),
    ("Resource ID", "masked_resource_id", 80),
]


def _sheet_resources(wb: Workbook, resources: list[AzureResource]) -> None:
    ws = wb.create_sheet("Resource Inventory")
    headers = [c[0] for c in _RESOURCES_COLS]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_idx, resource in enumerate(resources, start=2):
        alt = row_idx % 2 == 0
        for col_idx, (_, field, _) in enumerate(_RESOURCES_COLS, start=1):
            if field == "masked_resource_id":
                val: Any = resource.masked_resource_id()
            elif field == "masked_subscription_id":
                val = resource.masked_subscription_id()
            else:
                val = getattr(resource, field, None) or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if alt:
                cell.fill = _ALT_FILL

    for col_idx, (_, _, width) in enumerate(_RESOURCES_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    if resources:
        _add_table(ws, "TblInventory_ARG", len(resources))


# Sheet 12: Collection Metadata
# ---------------------------------------------------------------------------

def _sheet_metadata(wb: Workbook, metadata: CollectionMetadata) -> None:
    ws = wb.create_sheet("Run Metadata")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 50

    rows: list[tuple[str, Any]] = [
        ("Run Date (UTC)", metadata.run_date),
        ("Tool Version", metadata.tool_version),
        ("Metrics Period (days)", metadata.metrics_period_days),
        ("Total VM Count", metadata.total_vm_count),
        ("", ""),
        ("Subscriptions Scanned", ""),
    ]
    for sub_id in metadata.subscriptions_scanned:
        rows.append(("", sub_id))

    rows += [
        ("", ""),
        ("Thresholds Used", ""),
        ("  Underutilized CPU (avg %)", metadata.thresholds.underutilized_cpu_avg),
        ("  Underutilized Memory (avg %)", metadata.thresholds.underutilized_memory_avg),
        ("  Oversized CPU (P95 %)", metadata.thresholds.oversize_cpu_p95),
        ("  Headroom Multiplier", metadata.thresholds.headroom_multiplier),
        ("  PaaS Candidate CPU (avg %)", metadata.thresholds.paas_candidate_cpu_avg),
    ]

    for row_idx, (key, val) in enumerate(rows, start=1):
        key_cell = ws.cell(row=row_idx, column=1, value=key)
        val_cell = ws.cell(row=row_idx, column=2, value=val)
        if key and not key.startswith("  ") and val == "":
            key_cell.font = Font(bold=True, size=10)
        else:
            key_cell.font = Font(size=9)
            val_cell.font = Font(size=9)

    # ------------------------------------------------------------------
    # Performance Metrics Reference table
    # Checked metrics (✅) are the ones actively collected and used by
    # the recommendations engine.
    # ------------------------------------------------------------------
    _METRICS: list[tuple[str, str, bool]] = [
        ("Avg",          "baseline context",            False),
        ("P50",          "typical workload",            False),
        ("P95",          "primary optimization signal", True),
        ("P99",          "safety / risk signal",        True),
        ("Max",          "spike detection",             False),
        ("Min",          "anomaly / idle",              False),
        ("Data Points",  "confidence level",            True),
    ]
    _CHECKED_FILL = PatternFill("solid", fgColor="C6EFCE")   # light green

    # Blank separator row, then section header
    section_start = row_idx + 2
    hdr_cell = ws.cell(row=section_start, column=1, value="Performance Metrics Reference")
    hdr_cell.font = Font(bold=True, size=10)

    # Column headers
    col_hdr_row = section_start + 1
    for col, label in enumerate(("Metric", "Why"), start=1):
        cell = ws.cell(row=col_hdr_row, column=col, value=label)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="left")

    # Data rows
    for offset, (metric, reason, checked) in enumerate(_METRICS, start=1):
        data_row = col_hdr_row + offset
        label = f"\u2705 {metric}" if checked else metric
        m_cell = ws.cell(row=data_row, column=1, value=label)
        r_cell = ws.cell(row=data_row, column=2, value=reason)
        if checked:
            m_cell.fill = _CHECKED_FILL
            r_cell.fill = _CHECKED_FILL
        m_cell.font = Font(size=9, bold=checked)
        r_cell.font = Font(size=9)


# ---------------------------------------------------------------------------
# Sheet: Capacity Reservations (CRGs — counts and fill rates only)
# No $ / cost columns anywhere (SPEC §1.2).
# ---------------------------------------------------------------------------

_CRG_COLS: list[tuple[str, int]] = [
    ("Group Name",             28),
    ("Subscription (masked)",  36),
    ("Region",                 18),
    ("Zones",                  14),
    ("Reservation Name",       28),
    ("SKU",                    22),
    ("Reserved Count",         16),
    ("Used Count",             14),
    ("Fill Rate (%)",          14),
]


def _sheet_capacity_reservations(
    wb: Workbook,
    capacity_reservations: list[CapacityReservationGroup],
) -> None:
    ws = wb.create_sheet("Capacity Reservations")

    # Header row
    for col_idx, (label, _) in enumerate(_CRG_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="left")
        cell.border = _THIN_BORDER

    row_idx = 2
    for crg in capacity_reservations:
        zones_str = ", ".join(crg.zones) if crg.zones else ""
        masked_sub = crg.masked_subscription_id()
        if crg.reservations:
            for item in crg.reservations:
                fill_pct: float | None = None
                if item.reserved_count > 0:
                    fill_pct = round(100.0 * item.used_count / item.reserved_count, 1)
                vals: list[Any] = [
                    crg.group_name,
                    masked_sub,
                    crg.region,
                    zones_str,
                    item.reservation_name,
                    item.sku_name,
                    item.reserved_count,
                    item.used_count,
                    fill_pct,
                ]
                alt = row_idx % 2 == 0
                for col_idx, val in enumerate(vals, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = _THIN_BORDER
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(vertical="top", wrap_text=False)
                    if alt:
                        cell.fill = _ALT_FILL
                row_idx += 1
        else:
            # CRG with no reservations — emit one summary row
            vals = [
                crg.group_name, masked_sub, crg.region, zones_str,
                "", "", 0, 0, None,
            ]
            alt = row_idx % 2 == 0
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = _THIN_BORDER
                cell.font = Font(size=9)
                if alt:
                    cell.fill = _ALT_FILL
            row_idx += 1

    for col_idx, (_, width) in enumerate(_CRG_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    last_data_row = row_idx - 2  # last filled row
    if last_data_row >= 1:
        _add_table(ws, "TblCapacityReservations", last_data_row)


# ---------------------------------------------------------------------------
# Sheet: Deployment Failures (SPEC §3.5 / §11.11)
# ---------------------------------------------------------------------------

_DEPFAIL_COLS: list[tuple[str, int]] = [
    ("Resource Name",           28),
    ("Resource Type",           36),
    ("Subscription (masked)",   42),
    ("Resource Group",          22),
    ("Region",                  16),
    ("Error Class",             16),
    ("Operation Name",          40),
    ("Timestamp",               22),
    ("Status Message",          60),
]


def _sheet_deployment_failures(
    wb: Workbook,
    deployment_failures: list[DeploymentFailureEntry],
) -> None:
    ws = wb.create_sheet("Deployment Failures")

    for col_idx, (label, _) in enumerate(_DEPFAIL_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="left")
        cell.border = _THIN_BORDER

    row_idx = 2
    for f in deployment_failures:
        vals: list[Any] = [
            f.resource_name,
            f.resource_type,
            f.masked_subscription_id(),
            f.resource_group,
            f.region,
            f.error_class,
            f.operation_name,
            f.timestamp,
            f.status_message,
        ]
        alt = row_idx % 2 == 0
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if alt:
                cell.fill = _ALT_FILL
        row_idx += 1

    for col_idx, (_, width) in enumerate(_DEPFAIL_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    last_data_row = row_idx - 2
    if last_data_row >= 1:
        _add_table(ws, "TblDeploymentFailures", last_data_row)


# ---------------------------------------------------------------------------
# Read-back helpers (for export command)
# ---------------------------------------------------------------------------

def _read_inventory_sheet(wb) -> list[VmInventory]:
    # "Fleet Inventory" is the canonical VM-only sheet written since the excel-restructure.
    # "VM Inventory" is the legacy name from earlier builds.
    # "Resource Inventory" intentionally excluded — it holds ALL ARM resource types and
    # does not contain VM-specific columns (vCPUs, Memory, Power State etc.).
    sheet_name = next(
        (n for n in ("Fleet Inventory", "VM Inventory") if n in wb.sheetnames), None
    )
    if sheet_name is None:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]

    def col(row, name):
        try:
            return row[headers.index(name)]
        except (ValueError, IndexError):
            return None

    vms = []
    for row in rows[1:]:
        if not any(row):
            continue
        try:
            vms.append(
                VmInventory(
                    resource_id=col(row, "Resource ID") or "",
                    subscription_id="",  # masked in file, not recoverable
                    subscription_name=col(row, "Subscription") or "",
                    resource_group=col(row, "Resource Group") or "",
                    vm_name=col(row, "VM Name") or "",
                    vm_sku=col(row, "VM SKU") or "",
                    vcpus=int(col(row, "vCPUs") or 0),
                    memory_gb=float(col(row, "Memory (GB)") or 0),
                    region=col(row, "Region") or "",
                    os_type=col(row, "OS Type") or "Unknown",
                    os_version=col(row, "OS Version"),
                    power_state=col(row, "Power State"),
                    image_publisher=col(row, "Image Publisher"),
                    image_offer=col(row, "Image Offer"),
                    image_sku=col(row, "Image SKU"),
                    image_version=col(row, "Image Version"),
                    availability_zone=col(row, "Avail. Zone"),
                    nic_count=int(col(row, "NIC Count") or 0),
                    disk_count=int(col(row, "Disk Count") or 0),
                    disk_sizes_gb=[
                        float(s.strip()) for s in str(col(row, "Disk Sizes (GB)") or "").split(",")
                        if s.strip()
                    ],
                    vmss_name=col(row, "VMSS Name"),
                    availability_set_name=col(row, "Availability Set"),
                    workload=col(row, "Workload"),
                    application=col(row, "Application"),
                    environment=col(row, "Environment"),
                    criticality=col(row, "Criticality"),
                    owner=col(row, "Owner"),
                    custom=col(row, "Custom"),
                )
            )
        except Exception:
            continue
    return vms


def _read_raw_metrics_sheet(wb) -> list[VmMetrics]:
    """Read the Raw Metrics sheet back into VmMetrics objects (no time series).

    Returns an empty list when the Raw Metrics sheet is absent (dropped in Phase B).
    """
    if "Raw Metrics" not in wb.sheetnames:
        return []
    ws = wb["Raw Metrics"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    metrics: list[VmMetrics] = []
    for row in rows[1:]:
        if not any(row):
            continue
        try:
            metrics.append(VmMetrics(
                resource_id=str(row[0] or ""),
                metric_name=str(row[1] or ""),
                avg=float(row[2]) if row[2] is not None else None,
                p50=float(row[3]) if row[3] is not None else None,
                p95=float(row[4]) if row[4] is not None else None,
                p99=float(row[5]) if row[5] is not None else None,
                max=float(row[6]) if row[6] is not None else None,
                min=float(row[7]) if row[7] is not None else None,
                time_series=[],
            ))
        except Exception:
            continue
    return metrics


def _read_recommendations_sheet(wb) -> list[VmRecommendation]:
    # Support both Phase B "Decisions" table and legacy "Optimizations" sheet.
    # Phase B "Decisions" uses Finding fields; map the overlapping columns for
    # backward-compatible dashboard loading (best-effort).
    if "Optimizations" not in wb.sheetnames and "Decisions" not in wb.sheetnames:
        return []
    ws = wb["Optimizations"] if "Optimizations" in wb.sheetnames else wb["Decisions"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    recs = []
    for row in rows[1:]:
        if not any(row):
            continue
        try:
            # Column order:
            # 0 Priority | 1 Recommendation | 2 Category | 3 ResourceID
            # 4 Current SKU/RT | 5 Recommended SKU/RT | 6 Reason
            # 7 Estimated Optimization | 8 Override | 9 Notes
            current_val = str(row[4] or "")
            recommended_val = str(row[5] or "")
            recs.append(
                VmRecommendation(
                    priority=str(row[0] or "medium"),
                    recommendation=str(row[1] or ""),
                    category=str(row[2] or ""),
                    resource_id=str(row[3] or ""),
                    current_sku=current_val,
                    recommended_sku=recommended_val or None,
                    current_resource_type=current_val,
                    recommended_resource_type=recommended_val,
                    reason=str(row[6] or ""),
                    estimated_optimization=str(row[7] or ""),
                    manual_override=str(row[8]) if row[8] else None,
                    notes=str(row[9]) if row[9] else None,
                )
            )
        except Exception:
            continue
    return recs


def _read_metadata_sheet(wb) -> CollectionMetadata:
    from cloudopt.models import CollectionMetadata, CollectionThresholds

    # Support both Phase B "Run Metadata" and legacy "Collection Metadata"
    sheet_name = next(
        (n for n in ("Run Metadata", "Collection Metadata") if n in wb.sheetnames), None
    )
    if sheet_name is None:
        return CollectionMetadata(
            run_date="", tool_version="", subscriptions_scanned=[],
            metrics_period_days=30, total_vm_count=0,
            thresholds=CollectionThresholds(),
        )
    ws = wb[sheet_name]
    kv: dict[str, Any] = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] and row[1] is not None:
            kv[str(row[0]).strip()] = row[1]

    return CollectionMetadata(
        run_date=str(kv.get("Run Date (UTC)", "")),
        tool_version=str(kv.get("Tool Version", "")),
        subscriptions_scanned=[],
        metrics_period_days=int(kv.get("Metrics Period (days)", 30)),
        total_vm_count=int(kv.get("Total VM Count", 0)),
        thresholds=CollectionThresholds(
            underutilized_cpu_avg=float(kv.get("  Underutilized CPU (avg %)", 15)),
            underutilized_memory_avg=float(kv.get("  Underutilized Memory (avg %)", 20)),
            oversize_cpu_p95=float(kv.get("  Oversized CPU (P95 %)", 40)),
            headroom_multiplier=float(kv.get("  Headroom Multiplier", 1.2)),
            paas_candidate_cpu_avg=float(kv.get("  PaaS Candidate CPU (avg %)", 10)),
        ),
    )


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.row_dimensions[1].height = 18
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER


def _add_table(ws: Worksheet, name: str, data_rows: int) -> None:
    if data_rows < 1:
        return
    max_col = get_column_letter(ws.max_column)
    ref = f"A1:{max_col}{data_rows + 1}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    # A Table provides its own AutoFilter on the header row; the sheet-level
    # AutoFilter (set earlier via ws.auto_filter.ref) would produce a duplicate
    # <autoFilter> element that Excel rejects, causing tables to be removed.
    ws.auto_filter.ref = None


def _auto_width(ws: Worksheet, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 4)


def _colour_util(cell, value: float | None) -> None:
    if value is None:
        return
    if value < 40:
        cell.fill = _GREEN_FILL
    elif value < 70:
        cell.fill = _YELLOW_FILL
    else:
        cell.fill = _RED_FILL


def _fmt(val: float | None, decimals: int = 2) -> float | None:
    if val is None:
        return None
    return round(val, decimals)


def _group_metrics(metrics: list[VmMetrics]) -> dict[str, dict[str, VmMetrics]]:
    """Group metrics by resource_id → metric_name → VmMetrics."""
    result: dict[str, dict[str, VmMetrics]] = {}
    for m in metrics:
        result.setdefault(m.resource_id, {})[m.metric_name] = m
    return result


def _avg_metric(
    vms: list[VmInventory],
    metrics_by_vm: dict[str, dict[str, VmMetrics]],
    metric_name: str,
    stat: str,
) -> float | None:
    values = []
    for vm in vms:
        if _is_stopped(vm):
            continue
        m = metrics_by_vm.get(vm.resource_id, {}).get(metric_name)
        v = getattr(m, stat, None) if m else None
        if v is not None:
            values.append(v)
    if not values:
        return None
    return round(sum(values) / len(values), 2)


def _avg_mem_pct(
    vms: list[VmInventory],
    metrics_by_vm: dict[str, dict[str, VmMetrics]],
) -> float | None:
    """Return average memory utilization % across running VMs (derived from available bytes + total GB)."""
    values = []
    for vm in vms:
        if _is_stopped(vm):
            continue
        m = metrics_by_vm.get(vm.resource_id, {}).get("Available Memory Bytes")
        if m and m.avg is not None and vm.memory_gb > 0:
            avail_gb = m.avg / (1024 ** 3)
            used_pct = (1 - avail_gb / vm.memory_gb) * 100
            values.append(max(0.0, min(100.0, used_pct)))
    if not values:
        return None
    return round(sum(values) / len(values), 2)


def _is_stopped(vm: VmInventory) -> bool:
    """Return True when the VM is deallocated or stopped (not running)."""
    state = (vm.power_state or "").lower()
    return state in ("powerstate/stopped", "powerstate/deallocated")
