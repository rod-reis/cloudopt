"""Performance sheet writers for the cloudopt Excel workbook (SPEC §7.2).

Sheets produced:
  6.  Perf by VM - Standalone   — per-VM platform + guest metrics (standalone VMs only)
  7.  Perf by VM Group per SKU  — aggregated managed-service group rows
  8.  Perf by VM SKU per Subscription
  9.  Perf by VM SKU per Resource Group
"""

from __future__ import annotations

from statistics import mean
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from cloudopt.enrichment.schema import (
    CANONICAL_TO_GUEST_FIELD,
    EnrichedVmMetrics,
    GuestMetricRow,
)
from cloudopt.models import ManagedComputeGroupRow, VmInventory, VmMetrics

# Re-use shared helpers from the main excel module
from cloudopt.export.excel import (
    _ALT_FILL,
    _HDR_FILL,
    _HDR_FONT,
    _THIN_BORDER,
    _GREEN_FILL,
    _YELLOW_FILL,
    _RED_FILL,
    _write_header,
    _add_table,
    _colour_util,
    _avg_metric as _avg_metric_helper,
    _avg_mem_pct,
)

# ---------------------------------------------------------------------------
# SPEC §7.4 — ordered guest metric columns (39 columns)
# Each tuple: (Excel header label, GuestMetricRow field name)
# ---------------------------------------------------------------------------
_GUEST_COLS: list[tuple[str, str]] = [
    # OS group
    ("OS CPU Used %",          "os_cpu_used_percent"),
    ("OS Mem Used %",          "os_memory_used_percent"),
    ("OS Mem Avail (MB)",      "os_memory_available_mb"),
    ("OS Swap Used %",         "os_memory_swap_used_percent"),
    ("OS Disk Read IOps",      "os_disk_read_iops"),
    ("OS Disk Write IOps",     "os_disk_write_iops"),
    ("OS Disk Read MBps",      "os_disk_read_mbps"),
    ("OS Disk Write MBps",     "os_disk_write_mbps"),
    ("OS Net Recv MBps",       "os_network_receive_mbps"),
    ("OS Net Send MBps",       "os_network_send_mbps"),
    # JVM group
    ("JVM Heap Used %",        "jvm_heap_used_percent"),
    ("JVM Heap Used (MB)",     "jvm_heap_used_mb"),
    ("JVM Heap Max (MB)",      "jvm_heap_max_mb"),
    ("JVM GC Pause Avg (ms)",  "jvm_gc_pause_ms_avg"),
    ("JVM GC Pause P99 (ms)",  "jvm_gc_pause_ms_p99"),
    ("JVM Threads",            "jvm_threads_live_count"),
    # .NET group
    (".NET GC Heap (MB)",      "dotnet_gc_heap_size_mb"),
    (".NET GC Pause Avg (ms)", "dotnet_gc_pause_ms_avg"),
    (".NET TP Queue",          "dotnet_threadpool_queue_depth_avg"),
    (".NET Exceptions/s",      "dotnet_exceptions_rate"),
    # IIS group
    ("IIS Req/s",              "iis_requests_per_sec"),
    ("IIS Connections",        "iis_connections_current"),
    ("IIS Queue",              "iis_queue_length"),
    ("IIS Worker Restarts",    "iis_worker_restarts"),
    # SQL Server group
    ("SQL CPU %",              "sql_cpu_used_percent"),
    ("SQL Buf Pool Hit %",     "sql_memory_buffer_pool_hit_percent"),
    ("SQL Disk Read IOps",     "sql_disk_read_iops"),
    ("SQL Disk Write IOps",    "sql_disk_write_iops"),
    ("SQL Connections",        "sql_connections_active_count"),
    ("SQL Wait Avg (ms)",      "sql_waits_total_wait_ms_avg"),
    ("SQL Batch Req/s",        "sql_batch_requests_per_sec"),
    # PostgreSQL group
    ("PG CPU %",               "postgres_cpu_used_percent"),
    ("PG Connections",         "postgres_connections_active_count"),
    ("PG Cache Hit %",         "postgres_cache_hit_ratio_percent"),
    ("PG TPS",                 "postgres_transactions_per_sec"),
    # MySQL group
    ("MySQL CPU %",            "mysql_cpu_used_percent"),
    ("MySQL Connections",      "mysql_connections_active_count"),
    ("MySQL InnoDB Hit %",     "mysql_innodb_buffer_pool_hit_ratio_percent"),
    ("MySQL QPS",              "mysql_queries_per_sec"),
]


# ---------------------------------------------------------------------------
# Static columns for per-VM sheet
# ---------------------------------------------------------------------------
_VM_STATIC_COLS: list[tuple[str, str, int]] = [
    ("VM Name",          "vm_name",           28),
    ("Subscription",     "subscription_name", 26),
    ("Resource Group",   "resource_group",    24),
    ("Region",           "region",            16),
    ("VM SKU",           "vm_sku",            22),
    ("vCPUs",            "vcpus",             8),
    ("Memory (GB)",      "memory_gb",         12),
    ("OS Type",          "os_type",           12),
    ("Avail. Zone",      "availability_zone", 12),
    ("VMSS/AvSet Name", "__vmss_or_avset__",  22),
    ("Power State",      "power_state",       16),
]

# Platform metric columns — (header_label, azure_metric_name, stat_attr on VmMetrics)
_PLATFORM_COLS: list[tuple[str, str, str]] = [
    ("Avg CPU %",  "Percentage CPU", "avg"),
    ("P95 CPU %",  "Percentage CPU", "p95"),
    ("P99 CPU %",  "Percentage CPU", "p99"),
    ("Max CPU %",  "Percentage CPU", "max"),
    ("Min CPU %",  "Percentage CPU", "min"),
]
_MEM_HEADER = "Avg Mem %"  # computed separately (bytes → %)


def _build_enriched_lookup(
    enriched_list: list[EnrichedVmMetrics],
) -> dict[str, GuestMetricRow]:
    """Build resource_id → GuestMetricRow from the enriched-metrics list."""
    result: dict[str, GuestMetricRow] = {}
    for enriched in enriched_list:
        kwargs: dict[str, Any] = {}
        for dp in enriched.data_points:
            field_name = CANONICAL_TO_GUEST_FIELD.get(dp.metric_name)
            if field_name is None:
                continue
            val = dp.avg_value
            if val is not None and dp.unit == "bytes" and field_name.endswith("_mb"):
                val = val / (1024 * 1024)
            kwargs[field_name] = val
        if kwargs:
            resource_id = getattr(enriched, "resource_id", None) or enriched.vm_name
            result[resource_id] = GuestMetricRow(**kwargs)
    return result


def sheet_perf_standalone(
    wb: Workbook,
    vms: list[VmInventory],
    metrics_by_vm: dict[str, dict[str, VmMetrics]],
    enriched_list: list[EnrichedVmMetrics],
) -> None:
    """Sheet 6: Perf by VM."""
    from cloudopt.models import VmMetrics as _VmMetrics

    ws = wb.create_sheet("Perf by VM")

    static_hdrs = [c[0] for c in _VM_STATIC_COLS]
    platform_hdrs = [c[0] for c in _PLATFORM_COLS] + [_MEM_HEADER]
    guest_hdrs = [label for label, _ in _GUEST_COLS]
    headers = static_hdrs + platform_hdrs + guest_hdrs
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    enriched_map = _build_enriched_lookup(enriched_list)

    # Also try vm_name lookup for backwards compat
    name_to_enriched = {
        e.vm_name.lower(): e for e in enriched_list
    }

    for row_idx, vm in enumerate(vms, start=2):
        alt = row_idx % 2 == 0
        row_vals: list[Any] = []

        # Static columns
        for _, field, _ in _VM_STATIC_COLS:
            if field == "__vmss_or_avset__":
                row_vals.append(vm.vmss_name or vm.availability_set_name or "")
            else:
                row_vals.append(getattr(vm, field, None) or "")

        # Platform metrics — look up each Azure metric by name
        vm_all_metrics = metrics_by_vm.get(vm.resource_id, {})

        def _plat(azure_metric: str, stat: str) -> float | None:
            m = vm_all_metrics.get(azure_metric)
            v = getattr(m, stat, None) if m else None
            return round(v, 2) if v is not None else None

        def _mem_pct() -> float | None:
            m = vm_all_metrics.get("Available Memory Bytes")
            if m is None or m.avg is None or not vm.memory_gb:
                return None
            avail_gb = m.avg / (1024 ** 3)
            return round(100.0 * (1 - avail_gb / vm.memory_gb), 2)

        for _, azure_metric, stat in _PLATFORM_COLS:
            row_vals.append(_plat(azure_metric, stat))
        row_vals.append(_mem_pct())

        # Guest metrics
        guest = (
            enriched_map.get(vm.resource_id)
            or _build_enriched_lookup(
                [e for e in enriched_list if e.vm_name.lower() == vm.vm_name.lower()]
            ).get(vm.vm_name.lower())
        )
        for _, field_name in _GUEST_COLS:
            val = getattr(guest, field_name, None) if guest else None
            row_vals.append(val)

        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            if alt:
                cell.fill = _ALT_FILL
            # Colour CPU columns only (Avg/P95/P99/Max/Min CPU)
            cpu_start = len(_VM_STATIC_COLS) + 1
            cpu_end = cpu_start + len(_PLATFORM_COLS) - 1
            if cpu_start <= col_idx <= cpu_end and isinstance(val, (int, float)):
                _colour_util(cell, val)

    # Column widths
    all_widths = (
        [c[2] for c in _VM_STATIC_COLS]
        + [13] * (len(_PLATFORM_COLS) + 1)  # CPU cols + mem col
        + [16] * len(_GUEST_COLS)
    )
    for col_idx, w in enumerate(all_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    if vms:
        _add_table(ws, "TblPerfStandalone", len(vms))


# ---------------------------------------------------------------------------
# Group row columns (SPEC §7.3)
# ---------------------------------------------------------------------------
_GROUP_STATIC_COLS: list[tuple[str, str, int]] = [
    ("Parent ResourceType",         "parent_resource_type",   24),
    ("Parent ResourceName",         "parent_service_name",    28),
    ("Parent Pool/Node Group Name", "parent_pool_name",       22),
    ("VMSS Name",        "vmss_name",              22),
    ("VM SKU",           "vm_sku",                 22),
    ("Instance Count",   "instance_count",         14),
    ("Subscription",     "subscription_name",      26),
    ("Resource Group",   "resource_group",         24),
    ("Region",           "region",                 16),
    ("OS Type",          "os_type",                12),
    ("vCPUs",            "vcpus",                  8),
    ("Memory (GB)",      "memory_gb",              12),
    ("Zones",            "zones",                  12),
    ("Has OS Data",      "has_os_data",            12),
]

_GROUP_PLATFORM_COLS: list[tuple[str, str]] = [
    ("Avg CPU %",    "avg_cpu_pct"),
    ("P95 CPU %",    "p95_cpu_pct"),
    ("P99 CPU %",    "p99_cpu_pct"),
    ("Max CPU %",    "max_cpu_pct"),
    ("Min CPU %",    "min_cpu_pct"),
    ("Avg Mem %",    "avg_mem_pct"),
]


def sheet_perf_group_by_sku(
    wb: Workbook,
    groups: list[ManagedComputeGroupRow],
) -> None:
    """Sheet: Perf by VM Group — all VMSS entities (standalone + managed services)."""
    ws = wb.create_sheet("Perf by VM Group")

    static_hdrs = [c[0] for c in _GROUP_STATIC_COLS]
    platform_hdrs = [c[0] for c in _GROUP_PLATFORM_COLS]
    guest_hdrs = [label for label, _ in _GUEST_COLS]
    headers = static_hdrs + platform_hdrs + guest_hdrs
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_idx, grp in enumerate(groups, start=2):
        alt = row_idx % 2 == 0
        row_vals: list[Any] = []

        for _, field, _ in _GROUP_STATIC_COLS:
            val = getattr(grp, field, None)
            if hasattr(val, "value"):
                val = val.value
            row_vals.append(val if val is not None else "")

        for _, field in _GROUP_PLATFORM_COLS:
            row_vals.append(getattr(grp, field, None))

        guest_data = grp.guest_metrics or {}
        for _, field_name in _GUEST_COLS:
            row_vals.append(guest_data.get(field_name))

        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            if alt:
                cell.fill = _ALT_FILL
            cpu_start = len(_GROUP_STATIC_COLS) + 1
            cpu_end = cpu_start + 4  # Avg/P95/P99/Max/Min CPU
            if cpu_start <= col_idx <= cpu_end and isinstance(val, (int, float)):
                _colour_util(cell, val)

    all_widths = (
        [c[2] for c in _GROUP_STATIC_COLS]
        + [13] * len(_GROUP_PLATFORM_COLS)
        + [16] * len(_GUEST_COLS)
    )
    for col_idx, w in enumerate(all_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    if groups:
        _add_table(ws, "TblPerfVmGroup", len(groups))


def sheet_sku_by_subscription(
    wb: Workbook,
    vms: list[VmInventory],
    metrics_by_vm: dict[str, dict[str, VmMetrics]],
) -> None:
    """Sheet 8: Perf by VM SKU per Subscription."""
    _sheet_sku_flat(
        wb, vms, metrics_by_vm, "Perf by VM SKU per Subscription",
        ["subscription_name"], ["Subscription"], [30], "TblSkuBySub",
    )


def sheet_sku_by_resource_group(
    wb: Workbook,
    vms: list[VmInventory],
    metrics_by_vm: dict[str, dict[str, VmMetrics]],
) -> None:
    """Sheet 9: Perf by VM SKU per Resource Group."""
    _sheet_sku_flat(
        wb, vms, metrics_by_vm, "Perf by VM SKU - Resource Group",
        ["subscription_name", "resource_group"], ["Subscription", "Resource Group"],
        [30, 28], "TblSkuByRG",
    )


# ---------------------------------------------------------------------------
# Internal: flat SKU aggregation (used for sheets 8 & 9)
# ---------------------------------------------------------------------------

def _sheet_sku_flat(
    wb: Workbook,
    vms: list[VmInventory],
    metrics_by_vm: dict[str, dict[str, VmMetrics]],
    sheet_name: str,
    group_model_fields: list[str],
    col_headers: list[str],
    col_widths: list[int],
    table_name: str,
) -> None:
    """One row per (group_fields..., vm_sku) combination with averaged metrics."""
    ws = wb.create_sheet(sheet_name)
    headers = col_headers + [
        "VM SKU", "VM Count",
        "Avg CPU %", "P95 CPU %", "P99 CPU %", "Max CPU %", "Min CPU %", "Avg Mem %",
    ]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    row_groups: dict[tuple, list[VmInventory]] = {}
    for vm in vms:
        key = tuple(
            getattr(vm, f, None) or "(unknown)" for f in group_model_fields
        ) + (vm.vm_sku,)
        row_groups.setdefault(key, []).append(vm)

    n_group = len(group_model_fields)
    avg_cpu_col = n_group + 3

    for row_idx, (key, group_vms) in enumerate(sorted(row_groups.items()), start=2):
        alt = row_idx % 2 == 0
        *group_vals, sku = key

        def _avg(azure_metric: str, stat: str) -> float | None:
            vals = []
            for v in group_vms:
                m = metrics_by_vm.get(v.resource_id, {}).get(azure_metric)
                s = getattr(m, stat, None) if m else None
                if s is not None:
                    vals.append(s)
            return round(mean(vals), 2) if vals else None

        mem_vals = []
        for v in group_vms:
            m = metrics_by_vm.get(v.resource_id, {}).get("Available Memory Bytes")
            if m is None or m.avg is None or not v.memory_gb:
                continue
            avail_gb = m.avg / (1024 ** 3)
            mem_vals.append(100.0 * (1 - avail_gb / v.memory_gb))
        avg_mem = round(mean(mem_vals), 2) if mem_vals else None

        row_data = list(group_vals) + [
            sku, len(group_vms),
            _avg("Percentage CPU", "avg"),
            _avg("Percentage CPU", "p95"),
            _avg("Percentage CPU", "p99"),
            _avg("Percentage CPU", "max"),
            _avg("Percentage CPU", "min"),
            avg_mem,
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            if alt:
                cell.fill = _ALT_FILL
        for cpu_col in range(avg_cpu_col, avg_cpu_col + 5):
            _colour_util(ws.cell(row=row_idx, column=cpu_col), row_data[cpu_col - 1])
        _colour_util(ws.cell(row=row_idx, column=avg_cpu_col + 5), avg_mem)

    all_widths = col_widths + [22, 10, 13, 13, 13, 13, 13, 13]
    for col_idx, w in enumerate(all_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    if row_groups:
        _add_table(ws, table_name, len(row_groups))
